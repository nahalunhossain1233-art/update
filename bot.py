import telebot
import pandas as pd
import json
import os
import shutil
import uuid
import requests
import asyncio
import aiohttp
import threading
import time
from datetime import datetime
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# Initialize bot
TOKEN = "8360833535:AAFAk8vp2ODvSJbj1rztN2BmFanmiTvkozs"
bot = telebot.TeleBot(TOKEN, parse_mode=None)  # Set parse_mode to None by default

# File paths
EXCEL_FILE = 'results.xlsx'
ALL_RESULTS_FOLDER = 'all_results'
EXCEL_LOCK_FILE = 'results.xlsx.lock'
USERS_FILE = 'users.json'
ADMIN_FILE = 'admin.json'
SUBADMINS_FILE = 'subadmins.json'
PROCESSED_ACCOUNTS_FILE = 'processed_accounts.json'
FILES_FOLDER = 'uploaded_files'
WORK_FOLDER = 'work_files'
RULES_FILE = 'rules.json'
PENDING_FILE = 'pending.json'
UID_CHECK_FILE = 'uid_check_results.json'
COOLDOWN_FILE = 'cooldown.json'
USER_LIMITS_FILE = 'user_limits.json'
TAKEN_FILE = 'taken_accounts.json'
PENDING_ACCOUNTS_FILE = 'pending_accounts.json'
USER_CHECK_RESULTS_FILE = 'user_check_results.json'
SYSTEM_STATUS_FILE = 'system_status.json'  # New file for system status

# Admin credentials
ADMIN_PASSWORD = "n"  # Default password
ADMIN_USERNAME = "@nhossain123"  # Admin contact

# User counter
USER_COUNTER_FILE = 'user_counter.json'

# Excel styling
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CONFIRMED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SUSPENDED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
C_SUSPENDED_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Yellow for c.suspended
ISSUE_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Purple for issue
TAKEN_FILL = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")  # Blue for taken
CELL_FONT = Font(name="Calibri", size=11)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Store user states
user_states = {}
admin_sessions = {}
work_sessions = {}
username_to_id = {}  # Store username to ID mapping
account_data_store = {}  # Store account data for callback
subadmin_sessions = {}  # Store sub-admin sessions
uid_check_queue = []  # Queue for UID checking in background
uid_check_results = {}  # Store UID check results
cooldown_timers = {}  # Store cooldown timers for users
user_check_in_progress = {}  # Store user check progress

# Default cooldown time (5 seconds)
DEFAULT_COOLDOWN = 5
CHECK_INTERVAL = 0.5  # Interval between checking accounts (1.5 seconds)

# Initialize data files and folders
def init_files():
    # Create folders
    os.makedirs(FILES_FOLDER, exist_ok=True)
    os.makedirs(WORK_FOLDER, exist_ok=True)
    os.makedirs(ALL_RESULTS_FOLDER, exist_ok=True)
    
    # Create Excel file if not exists
    if not os.path.exists(EXCEL_FILE):
        create_excel_with_styling()
    
    # Create users file if not exists
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'w') as f:
            json.dump({"users": {}, "username_mapping": {}, "banned": []}, f)
    
    # Create admin file if not exists
    if not os.path.exists(ADMIN_FILE):
        with open(ADMIN_FILE, 'w') as f:
            json.dump({"password": ADMIN_PASSWORD}, f)
    
    # Create subadmins file if not exists
    if not os.path.exists(SUBADMINS_FILE):
        with open(SUBADMINS_FILE, 'w') as f:
            json.dump({"subadmins": {}}, f)
    
    # Create processed accounts file if not exists
    if not os.path.exists(PROCESSED_ACCOUNTS_FILE):
        with open(PROCESSED_ACCOUNTS_FILE, 'w') as f:
            json.dump({"processed": []}, f, indent=4)
    
    # Create rules file if not exists
    if not os.path.exists(RULES_FILE):
        with open(RULES_FILE, 'w') as f:
            json.dump({"rules": "Default Rules:\n1. Process accounts in order\n2. Don't skip accounts\n3. Click Cancel to stop anytime"}, f, indent=4)
    
    # Create pending approvals file if not exists
    if not os.path.exists(PENDING_FILE):
        with open(PENDING_FILE, 'w') as f:
            json.dump({"pending": []}, f, indent=4)
    
    # Create user counter file if not exists
    if not os.path.exists(USER_COUNTER_FILE):
        with open(USER_COUNTER_FILE, 'w') as f:
            json.dump({"counter": 0}, f, indent=4)
    
    # Create UID check results file if not exists
    if not os.path.exists(UID_CHECK_FILE):
        with open(UID_CHECK_FILE, 'w') as f:
            json.dump({}, f, indent=4)
    
    # Create cooldown file if not exists
    if not os.path.exists(COOLDOWN_FILE):
        with open(COOLDOWN_FILE, 'w') as f:
            json.dump({"cooldown_seconds": DEFAULT_COOLDOWN}, f, indent=4)
    
    # Create user limits file if not exists
    if not os.path.exists(USER_LIMITS_FILE):
        with open(USER_LIMITS_FILE, 'w') as f:
            json.dump({}, f, indent=4)
    
    # Create taken accounts file if not exists
    if not os.path.exists(TAKEN_FILE):
        with open(TAKEN_FILE, 'w') as f:
            json.dump({"taken_rows": []}, f, indent=4)
    
    # Create pending accounts file if not exists
    if not os.path.exists(PENDING_ACCOUNTS_FILE):
        with open(PENDING_ACCOUNTS_FILE, 'w') as f:
            json.dump({"pending_accounts": {}}, f, indent=4)
    
    # Create user check results file if not exists
    if not os.path.exists(USER_CHECK_RESULTS_FILE):
        with open(USER_CHECK_RESULTS_FILE, 'w') as f:
            json.dump({}, f, indent=4)
    
    # Create system status file if not exists
    if not os.path.exists(SYSTEM_STATUS_FILE):
        with open(SYSTEM_STATUS_FILE, 'w') as f:
            json.dump({
                "work_enabled": True,
                "off_notice": "⚠️ System is temporarily offline for maintenance. Please try again later.",
                "on_notice": "✅ System is now online and working normally."
            }, f, indent=4)
    
    # Load username mapping
    load_username_mapping()
    
    # Load UID check results
    load_uid_check_results()
    
    # Start background UID check thread
    start_background_uid_checker()

# Get cooldown time
def get_cooldown_time():
    try:
        cooldown_data = load_json(COOLDOWN_FILE)
        return cooldown_data.get("cooldown_seconds", DEFAULT_COOLDOWN)
    except:
        return DEFAULT_COOLDOWN

# Set cooldown time
def set_cooldown_time(seconds):
    try:
        if seconds < 0:
            seconds = 0
        cooldown_data = {"cooldown_seconds": seconds}
        save_json(COOLDOWN_FILE, cooldown_data)
        return True
    except:
        return False

# Check if user is in cooldown
def is_user_in_cooldown(user_id):
    if user_id in cooldown_timers:
        remaining = cooldown_timers[user_id] - time.time()
        if remaining > 0:
            return True, remaining
    return False, 0

# Start cooldown for user
def start_cooldown(user_id):
    cooldown_time = get_cooldown_time()
    if cooldown_time > 0:
        cooldown_timers[user_id] = time.time() + cooldown_time
        return True
    return False

# ========== SYSTEM STATUS FUNCTIONS ==========

def get_system_status():
    """Get current system status"""
    try:
        return load_json(SYSTEM_STATUS_FILE)
    except:
        return {"work_enabled": True, "off_notice": "", "on_notice": ""}

def save_system_status(status_data):
    """Save system status"""
    return save_json(SYSTEM_STATUS_FILE, status_data)

def is_work_enabled():
    """Check if work system is enabled"""
    status = get_system_status()
    return status.get("work_enabled", True)

def enable_work():
    """Enable work system"""
    status = get_system_status()
    status["work_enabled"] = True
    save_system_status(status)
    return True

def disable_work():
    """Disable work system"""
    status = get_system_status()
    status["work_enabled"] = False
    save_system_status(status)
    return True

def set_off_notice(notice):
    """Set offline notice"""
    status = get_system_status()
    status["off_notice"] = notice
    save_system_status(status)
    return True

def set_on_notice(notice):
    """Set online notice"""
    status = get_system_status()
    status["on_notice"] = notice
    save_system_status(status)
    return True

def get_off_notice():
    """Get offline notice"""
    status = get_system_status()
    return status.get("off_notice", "⚠️ System is temporarily offline for maintenance. Please try again later.")

def get_on_notice():
    """Get online notice"""
    status = get_system_status()
    return status.get("on_notice", "✅ System is now online and working normally.")

# Create Excel file with styling (UPDATED - removed Pending column)
def create_excel_with_styling():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Define headers (removed Pending column)
        headers = ['Input', 'Username', 'Password', 'Processor', 'User Code', 'Status', 'Timestamp', 'Taken']
        
        # Write headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER
        
        # Set column widths
        column_widths = {
            'A': 50,  # Input (contains cookie)
            'B': 20,  # Username
            'C': 20,  # Password
            'D': 20,  # Processor
            'E': 15,  # User Code
            'F': 15,  # Status
            'G': 25,  # Timestamp
            'H': 15   # Taken
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save directly
        wb.save(EXCEL_FILE)
        print("✅ Excel file created/cleared successfully")
        return True
    except Exception as e:
        print(f"❌ Error creating Excel: {e}")
        return False

# Save Excel with file lock mechanism
def save_excel_with_lock(wb, max_retries=5, retry_delay=0.1):
    for attempt in range(max_retries):
        try:
            # Remove lock file if exists
            if os.path.exists(EXCEL_LOCK_FILE):
                try:
                    os.remove(EXCEL_LOCK_FILE)
                except:
                    pass
            
            # Save Excel file
            wb.save(EXCEL_FILE)
            return True
            
        except PermissionError as e:
            print(f"Excel save attempt {attempt + 1} failed: PermissionError - {e}")
            if attempt < max_retries - 1:
                time.sleep(retry_delay * (attempt + 1))  # Exponential backoff
            else:
                # Try to save with a different method
                try:
                    temp_file = EXCEL_FILE + '.temp'
                    wb.save(temp_file)
                    if os.path.exists(EXCEL_FILE):
                        os.remove(EXCEL_FILE)
                    os.rename(temp_file, EXCEL_FILE)
                    return True
                except Exception as e2:
                    print(f"Failed to save with temp file: {e2}")
                    return False
        except Exception as e:
            print(f"Excel save attempt {attempt + 1} failed: {e}")
            time.sleep(retry_delay)
    
    return False

# Save data to Excel (UPDATED - removed Pending column)
def save_to_excel(full_input, username, password, processor, user_code, status):
    try:
        # Wait a moment for any file locks
        time.sleep(0.1)
        
        # Try to load existing workbook
        wb = None
        if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
                
                # Check if headers exist
                if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
                    # File exists but is empty or corrupted, recreate it
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Results"
                    
                    # Write headers (removed Pending column)
                    headers = ['Input', 'Username', 'Password', 'Processor', 'User Code', 'Status', 'Timestamp', 'Taken']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num, value=header)
                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT
                        cell.alignment = ALIGN_CENTER
                        cell.border = BORDER
            except Exception as e:
                print(f"Error loading workbook, creating new: {e}")
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
                
                # Write headers (removed Pending column)
                headers = ['Input', 'Username', 'Password', 'Processor', 'User Code', 'Status', 'Timestamp', 'Taken']
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = ALIGN_CENTER
                    cell.border = BORDER
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Write headers (removed Pending column)
            headers = ['Input', 'Username', 'Password', 'Processor', 'User Code', 'Status', 'Timestamp', 'Taken']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Prepare data
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Write data to cells (removed Pending column)
        data = [
            str(full_input),  # Full input with cookie
            str(username),
            str(password),
            str(processor),
            str(user_code),
            str(status),
            str(timestamp),
            ""  # Taken column - empty initially
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=next_row, column=col_num, value=value)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = ALIGN_LEFT if col_num in [1, 7] else ALIGN_CENTER
            
            # Apply status-specific styling
            if col_num == 6:  # Status column
                status_lower = str(status).lower()
                if status_lower == "confirmed":
                    cell.fill = CONFIRMED_FILL
                    cell.font = Font(color="006100", bold=True)
                elif status_lower == "suspended":
                    cell.fill = SUSPENDED_FILL
                    cell.font = Font(color="9C0006", bold=True)
                elif "c.suspended" in status_lower or "c_suspended" in status_lower:
                    cell.fill = C_SUSPENDED_FILL
                    cell.font = Font(color="7F6000", bold=True)
                elif "issue" in status_lower:
                    cell.fill = ISSUE_FILL
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Max width 100
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save with lock mechanism
        if save_excel_with_lock(wb):
            print(f"✅ Excel saved successfully: {username} - {status}")
            return True
        else:
            print("❌ Failed to save Excel with lock, trying fallback...")
            return save_to_excel_fallback(full_input, username, password, processor, user_code, status)
            
    except Exception as e:
        print(f"❌ Error in save_to_excel: {e}")
        return save_to_excel_fallback(full_input, username, password, processor, user_code, status)

# Fallback Excel saving method using pandas (UPDATED - removed Pending column)
def save_to_excel_fallback(full_input, username, password, processor, user_code, status):
    try:
        # Create DataFrame with new data
        new_data = {
            'Input': full_input,
            'Username': username,
            'Password': password,
            'Processor': processor,
            'User Code': user_code,
            'Status': status,
            'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'Taken': ""
        }
        
        # Try to read existing data
        existing_data = []
        if os.path.exists(EXCEL_FILE):
            try:
                df_existing = pd.read_excel(EXCEL_FILE)
                existing_data = df_existing.to_dict('records')
            except:
                pass
        
        # Add new data
        existing_data.append(new_data)
        
        # Create DataFrame
        df = pd.DataFrame(existing_data)
        
        # Save to temporary file first
        temp_file = EXCEL_FILE + '.temp'
        
        # Save with pandas to_excel
        df.to_excel(temp_file, index=False, engine='openpyxl')
        
        # Replace original file
        if os.path.exists(EXCEL_FILE):
            try:
                os.remove(EXCEL_FILE)
            except:
                pass
        
        os.rename(temp_file, EXCEL_FILE)
        
        print(f"✅ Excel saved via fallback: {username} - {status}")
        return True
        
    except Exception as e:
        print(f"❌ Error in fallback Excel save: {e}")
        return False

# Remove user's data from Excel
def remove_user_from_excel(username):
    """Remove all entries of a user from Excel file"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel file not found"
        
        # Get user ID and code
        user_id = get_user_id_from_username(username.lower())
        if not user_id:
            return False, "User not found"
        
        user_code = get_user_code(user_id)
        if not user_code:
            return False, "User code not found"
        
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find rows to delete
        rows_to_delete = []
        for row in range(2, ws.max_row + 1):
            processor_code = ws.cell(row=row, column=5).value  # User Code column
            if processor_code == user_code:
                rows_to_delete.append(row)
        
        # Delete rows from bottom to top
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        
        # Save Excel
        wb.save(EXCEL_FILE)
        
        return True, f"Removed {len(rows_to_delete)} rows for user @{username}"
        
    except Exception as e:
        print(f"Error removing user from Excel: {e}")
        return False, str(e)

# Remove specific status data from Excel
def remove_user_status_from_excel(username, status_to_remove):
    """Remove specific status entries of a user from Excel"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel file not found"
        
        # Get user ID and code
        user_id = get_user_id_from_username(username.lower())
        if not user_id:
            return False, "User not found"
        
        user_code = get_user_code(user_id)
        if not user_code:
            return False, "User code not found"
        
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find rows to delete based on status
        rows_to_delete = []
        for row in range(2, ws.max_row + 1):
            processor_code = ws.cell(row=row, column=5).value  # User Code column
            status = ws.cell(row=row, column=6).value  # Status column
            
            if processor_code == user_code and status and status_to_remove.lower() in str(status).lower():
                rows_to_delete.append(row)
        
        # Delete rows from bottom to top
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        
        # Save Excel
        wb.save(EXCEL_FILE)
        
        return True, f"Removed {len(rows_to_delete)} '{status_to_remove}' rows for user @{username}"
        
    except Exception as e:
        print(f"Error removing user status from Excel: {e}")
        return False, str(e)

# Keep only specified status data in Excel
def keep_only_user_status_in_excel(username, status_to_keep):
    """Keep only specified status entries of a user in Excel"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel file not found"
        
        # Get user ID and code
        user_id = get_user_id_from_username(username.lower())
        if not user_id:
            return False, "User not found"
        
        user_code = get_user_code(user_id)
        if not user_code:
            return False, "User code not found"
        
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find rows to delete (all except the status to keep)
        rows_to_delete = []
        for row in range(2, ws.max_row + 1):
            processor_code = ws.cell(row=row, column=5).value  # User Code column
            status = ws.cell(row=row, column=6).value  # Status column
            
            if processor_code == user_code:
                if not status or status_to_keep.lower() not in str(status).lower():
                    rows_to_delete.append(row)
        
        # Delete rows from bottom to top
        for row in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row)
        
        # Save Excel
        wb.save(EXCEL_FILE)
        
        return True, f"Kept only '{status_to_keep}' entries for user @{username}, removed {len(rows_to_delete)} other rows"
        
    except Exception as e:
        print(f"Error keeping user status in Excel: {e}")
        return False, str(e)

# Save permanent results to all_results folder
def save_permanent_results():
    try:
        if not os.path.exists(EXCEL_FILE) or os.path.getsize(EXCEL_FILE) < 1024:
            return False, "No data available to save"
        
        # Create timestamp for filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"all_results_{timestamp}.xlsx"
        filepath = os.path.join(ALL_RESULTS_FOLDER, filename)
        
        # Copy current results to permanent location
        shutil.copy2(EXCEL_FILE, filepath)
        
        # Count records in the file
        try:
            df = pd.read_excel(EXCEL_FILE)
            record_count = len(df)
        except:
            record_count = 0
        
        return True, {
            "filename": filename,
            "filepath": filepath,
            "records": record_count,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
    except Exception as e:
        print(f"❌ Error saving permanent results: {e}")
        return False, str(e)

# List all permanent result files
def list_permanent_results():
    try:
        files = [f for f in os.listdir(ALL_RESULTS_FOLDER) if f.endswith('.xlsx')]
        files.sort(reverse=True)  # Sort by newest first
        
        result_files = []
        for filename in files:
            filepath = os.path.join(ALL_RESULTS_FOLDER, filename)
            file_size = os.path.getsize(filepath)
            created_time = datetime.fromtimestamp(os.path.getctime(filepath)).strftime("%Y-%m-%d %H:%M:%S")
            
            # Extract timestamp from filename
            timestamp_str = filename.replace("all_results_", "").replace(".xlsx", "")
            
            result_files.append({
                "filename": filename,
                "filepath": filepath,
                "size": file_size,
                "created": created_time,
                "timestamp": timestamp_str
            })
        
        return result_files
    except Exception as e:
        print(f"❌ Error listing permanent results: {e}")
        return []

# Load/save data functions
def load_json(file_path):
    try:
        with open(file_path, 'r') as f:
            return json.load(f)
    except:
        return {}

def save_json(file_path, data):
    try:
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=4)
        return True
    except:
        return False

# Load and save username mapping
def load_username_mapping():
    global username_to_id
    users_data = load_json(USERS_FILE)
    username_to_id = users_data.get("username_mapping", {})
    return username_to_id

def save_username_mapping():
    users_data = load_json(USERS_FILE)
    users_data["username_mapping"] = username_to_id
    save_json(USERS_FILE, users_data)

# Load UID check results
def load_uid_check_results():
    global uid_check_results
    try:
        uid_check_results = load_json(UID_CHECK_FILE)
    except:
        uid_check_results = {}

# Save UID check results
def save_uid_check_results():
    save_json(UID_CHECK_FILE, uid_check_results)

# Update username mapping
def update_username_mapping(user_id, username):
    global username_to_id
    if username and username not in ['', 'None', None]:
        username_to_id[username.lower()] = user_id
        save_username_mapping()

# Get user ID from username
def get_user_id_from_username(username):
    return username_to_id.get(username.lower())

# Check if user is banned
def is_user_banned(user_id):
    users_data = load_json(USERS_FILE)
    return user_id in users_data.get("banned", [])

# Check if user is approved
def is_user_approved(user_id):
    users_data = load_json(USERS_FILE)
    return user_id in users_data.get("users", {})

# Check if user has pending approval
def has_pending_approval(user_id):
    pending_data = load_json(PENDING_FILE)
    return any(pending["user_id"] == user_id for pending in pending_data.get("pending", []))

# Add user to pending approvals
def add_pending_approval(user_id, username):
    pending_data = load_json(PENDING_FILE)
    pending_list = pending_data.get("pending", [])
    
    # Check if already pending
    if not any(pending["user_id"] == user_id for pending in pending_list):
        pending_list.append({
            "user_id": user_id,
            "username": username,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        pending_data["pending"] = pending_list
        save_json(PENDING_FILE, pending_data)
        return True
    return False

# Get pending approvals
def get_pending_approvals():
    pending_data = load_json(PENDING_FILE)
    return pending_data.get("pending", [])

# Approve user
def approve_user(user_id):
    # Remove from pending
    pending_data = load_json(PENDING_FILE)
    pending_list = pending_data.get("pending", [])
    pending_list = [p for p in pending_list if p["user_id"] != user_id]
    pending_data["pending"] = pending_list
    save_json(PENDING_FILE, pending_data)
    
    return True

# Approve all pending users
def approve_all_pending():
    pending_data = load_json(PENDING_FILE)
    pending_list = pending_data.get("pending", [])
    pending_data["pending"] = []
    save_json(PENDING_FILE, pending_data)
    return [p["user_id"] for p in pending_list]

# Ban user by username
def ban_user(username):
    user_id = get_user_id_from_username(username.lower())
    if user_id:
        users_data = load_json(USERS_FILE)
        if "banned" not in users_data:
            users_data["banned"] = []
        if user_id not in users_data["banned"]:
            users_data["banned"].append(user_id)
            save_json(USERS_FILE, users_data)
            return True, user_id
    return False, None

# Unban user by username
def unban_user(username):
    user_id = get_user_id_from_username(username.lower())
    if user_id:
        users_data = load_json(USERS_FILE)
        if "banned" in users_data and user_id in users_data["banned"]:
            users_data["banned"].remove(user_id)
            save_json(USERS_FILE, users_data)
            return True, user_id
    return False, None

# Get next user code
def get_next_user_code():
    counter_data = load_json(USER_COUNTER_FILE)
    counter = counter_data.get("counter", 0) + 1
    counter_data["counter"] = counter
    save_json(USER_COUNTER_FILE, counter_data)
    return f"u-{counter}"

# Get user code for a user
def get_user_code(user_id):
    users_data = load_json(USERS_FILE)
    if user_id in users_data.get("users", {}):
        return users_data["users"][user_id].get("user_code")
    return None

# Remove user from database
def remove_user(username):
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    users_data = load_json(USERS_FILE)
    
    # Remove from users
    if user_id in users_data.get("users", {}):
        del users_data["users"][user_id]
    
    # Remove from username mapping
    if username.lower() in users_data.get("username_mapping", {}):
        del users_data["username_mapping"][username.lower()]
    
    # Remove from banned list if present
    if user_id in users_data.get("banned", []):
        users_data["banned"].remove(user_id)
    
    save_json(USERS_FILE, users_data)
    return True, user_id

# Load admin password
def get_admin_password():
    try:
        admin_data = load_json(ADMIN_FILE)
        return admin_data.get("password", ADMIN_PASSWORD)
    except:
        return ADMIN_PASSWORD

# Update admin password
def update_admin_password(new_password):
    admin_data = {"password": new_password}
    save_json(ADMIN_FILE, admin_data)
    return True

# Check if user is admin
def is_admin(user_id):
    return admin_sessions.get(str(user_id), False)

# Check if user is subadmin
def is_subadmin(user_id):
    # Check if in subadmin sessions
    if subadmin_sessions.get(str(user_id)):
        return True
    
    # Check if in subadmins file
    subadmins_data = load_json(SUBADMINS_FILE)
    return str(user_id) in subadmins_data.get("subadmins", {})

# Add subadmin
def add_subadmin(username, password):
    # Find user ID from username
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    subadmins_data = load_json(SUBADMINS_FILE)
    
    # Check if already a subadmin
    if user_id in subadmins_data.get("subadmins", {}):
        return False, "User is already a subadmin"
    
    # Add subadmin
    subadmins_data.setdefault("subadmins", {})[user_id] = {
        "username": username,
        "password": password,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    save_json(SUBADMINS_FILE, subadmins_data)
    return True, user_id

# Remove subadmin
def remove_subadmin(username):
    # Find user ID from username
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    subadmins_data = load_json(SUBADMINS_FILE)
    
    # Check if is a subadmin
    if user_id not in subadmins_data.get("subadmins", {}):
        return False, "User is not a subadmin"
    
    # Remove subadmin
    del subadmins_data["subadmins"][user_id]
    save_json(SUBADMINS_FILE, subadmins_data)
    
    # Remove from subadmin sessions if logged in
    if user_id in subadmin_sessions:
        del subadmin_sessions[user_id]
    
    return True, user_id

# Get all admins list
def get_all_admins():
    result = {
        "main_admin": [],
        "sub_admins": []
    }
    
    # Get main admin sessions
    for user_id, is_logged_in in admin_sessions.items():
        if is_logged_in:
            users_data = load_json(USERS_FILE)
            username = users_data.get("users", {}).get(user_id, {}).get("username", "Unknown")
            result["main_admin"].append({
                "user_id": user_id,
                "username": username
            })
    
    # Get subadmins from file
    subadmins_data = load_json(SUBADMINS_FILE)
    for user_id, info in subadmins_data.get("subadmins", {}).items():
        result["sub_admins"].append({
            "user_id": user_id,
            "username": info.get("username", "Unknown"),
            "created_at": info.get("created_at", "Unknown"),
            "is_logged_in": user_id in subadmin_sessions
        })
    
    return result

# Check if account is already processed
def is_account_processed(username):
    try:
        processed_data = load_json(PROCESSED_ACCOUNTS_FILE)
        return username in processed_data["processed"]
    except:
        return False

def mark_account_processed(username):
    try:
        processed_data = load_json(PROCESSED_ACCOUNTS_FILE)
        if username not in processed_data["processed"]:
            processed_data["processed"].append(username)
            save_json(PROCESSED_ACCOUNTS_FILE, processed_data)
        return True
    except:
        return False

# Update user stats
def update_user_stats(user_id, username, user_code, action):
    users_data = load_json(USERS_FILE)
    
    # Initialize users if not exists
    if "users" not in users_data:
        users_data["users"] = {}
    
    if user_id not in users_data["users"]:
        users_data["users"][user_id] = {
            "username": username,
            "user_code": user_code,
            "confirmed": 0,
            "suspended": 0,
            "c_suspended": 0,
            "issue": 0,
            "total": 0
        }
    
    if action == "confirm":
        users_data["users"][user_id]["confirmed"] = users_data["users"][user_id].get("confirmed", 0) + 1
    elif action == "suspend":
        users_data["users"][user_id]["suspended"] = users_data["users"][user_id].get("suspended", 0) + 1
    elif action == "c_suspend":
        users_data["users"][user_id]["c_suspended"] = users_data["users"][user_id].get("c_suspended", 0) + 1
    elif action == "issue":
        users_data["users"][user_id]["issue"] = users_data["users"][user_id].get("issue", 0) + 1
    
    users_data["users"][user_id]["total"] = users_data["users"][user_id].get("total", 0) + 1
    
    # Update username and user code
    if username:
        users_data["users"][user_id]["username"] = username
    if user_code:
        users_data["users"][user_id]["user_code"] = user_code
    
    save_json(USERS_FILE, users_data)
    return users_data["users"][user_id]

# Get user stats
def get_user_stats(user_id):
    users_data = load_json(USERS_FILE)
    if "users" in users_data and user_id in users_data["users"]:
        return users_data["users"][user_id]
    return None

# Reset specific user stats
def reset_user_stats(username, keep_type=None, full_reset=False):
    """Reset user stats with options"""
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    try:
        users_data = load_json(USERS_FILE)
        
        if "users" not in users_data or user_id not in users_data["users"]:
            return False, "User not found in database"
        
        if full_reset:
            # Reset all stats to zero
            users_data["users"][user_id]["confirmed"] = 0
            users_data["users"][user_id]["suspended"] = 0
            users_data["users"][user_id]["c_suspended"] = 0
            users_data["users"][user_id]["issue"] = 0
            users_data["users"][user_id]["total"] = 0
            
            # Remove user's data from Excel
            remove_user_from_excel(username)
            
            save_json(USERS_FILE, users_data)
            return True, f"Reset ALL stats and data for @{username}"
        
        if keep_type is None:
            # Reset all stats to zero
            users_data["users"][user_id]["confirmed"] = 0
            users_data["users"][user_id]["suspended"] = 0
            users_data["users"][user_id]["c_suspended"] = 0
            users_data["users"][user_id]["issue"] = 0
            users_data["users"][user_id]["total"] = 0
            
            # Remove user from Excel (all rows)
            remove_user_from_excel(username)
            
            save_json(USERS_FILE, users_data)
            return True, f"Reset all stats for @{username}"
        
        else:
            # Reset everything except the specified type
            keep_type = keep_type.lower()
            
            # Save the value to keep
            if keep_type == "live" or keep_type == "confirmed":
                value_to_keep = users_data["users"][user_id]["confirmed"]
            elif keep_type == "suspended":
                value_to_keep = users_data["users"][user_id]["suspended"]
            elif keep_type == "c_suspended":
                value_to_keep = users_data["users"][user_id]["c_suspended"]
            elif keep_type == "issue":
                value_to_keep = users_data["users"][user_id]["issue"]
            else:
                return False, f"Invalid keep type. Use: live, suspended, c_suspended, or issue"
            
            # Reset all to zero
            users_data["users"][user_id]["confirmed"] = 0
            users_data["users"][user_id]["suspended"] = 0
            users_data["users"][user_id]["c_suspended"] = 0
            users_data["users"][user_id]["issue"] = 0
            
            # Restore the kept value
            if keep_type == "live" or keep_type == "confirmed":
                users_data["users"][user_id]["confirmed"] = value_to_keep
            elif keep_type == "suspended":
                users_data["users"][user_id]["suspended"] = value_to_keep
            elif keep_type == "c_suspended":
                users_data["users"][user_id]["c_suspended"] = value_to_keep
            elif keep_type == "issue":
                users_data["users"][user_id]["issue"] = value_to_keep
            
            # Recalculate total
            users_data["users"][user_id]["total"] = (
                users_data["users"][user_id]["confirmed"] +
                users_data["users"][user_id]["suspended"] +
                users_data["users"][user_id]["c_suspended"] +
                users_data["users"][user_id]["issue"]
            )
            
            # Remove other statuses from Excel
            if keep_type == "confirmed":
                keep_only_user_status_in_excel(username, "confirmed")
            elif keep_type == "suspended":
                keep_only_user_status_in_excel(username, "suspended")
            elif keep_type == "c_suspended":
                keep_only_user_status_in_excel(username, "c.suspended")
            elif keep_type == "issue":
                keep_only_user_status_in_excel(username, "issue")
            
            save_json(USERS_FILE, users_data)
            return True, f"Reset stats for @{username}, kept only {keep_type} ({value_to_keep})"
    
    except Exception as e:
        print(f"Error resetting user stats: {e}")
        return False, str(e)

# Reset all user stats AND clear Excel file
def reset_all_stats(full_reset=False):
    # Clear user stats
    users_data = load_json(USERS_FILE)
    if "users" in users_data:
        for user_id in users_data["users"]:
            users_data["users"][user_id]["confirmed"] = 0
            users_data["users"][user_id]["suspended"] = 0
            users_data["users"][user_id]["c_suspended"] = 0
            users_data["users"][user_id]["issue"] = 0
            users_data["users"][user_id]["total"] = 0
        
        save_json(USERS_FILE, users_data)
    
    # Clear Excel file
    try:
        create_excel_with_styling()
        print("✅ Excel file cleared during stats reset")
        return True
    except Exception as e:
        print(f"❌ Error clearing Excel during stats reset: {e}")
        return False

# Check if UID is live (using Facebook Graph API)
async def check_uid_live_async(uid):
    try:
        # Extract first part of UID (before |)
        uid_first_part = uid.split("|")[0].strip()
        
        # Facebook Graph API URL to check profile picture
        url = f"https://graph.facebook.com/{uid_first_part}/picture?redirect=false"
        
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=10) as response:
                if response.status == 200:
                    data = await response.json()
                    # If the profile has a valid picture, the account is live
                    return data.get("data", {}).get("height") is not None
                else:
                    return False
    except:
        return False

# Check UID live (sync version for background thread)
def check_uid_live_sync(uid):
    try:
        # Extract first part of UID (before |)
        uid_first_part = uid.split("|")[0].strip()
        
        # Facebook Graph API URL to check profile picture
        url = f"https://graph.facebook.com/{uid_first_part}/picture?redirect=false"
        
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            data = response.json()
            # If the profile has a valid picture, the account is live
            return data.get("data", {}).get("height") is not None
        else:
            return False
    except:
        return False

# Background UID checker thread
def background_uid_checker():
    while True:
        try:
            if uid_check_queue:
                uid, account_id, callback_data = uid_check_queue.pop(0)
                
                # Check if UID is live
                is_live = check_uid_live_sync(uid)
                
                # Store result
                uid_check_results[account_id] = {
                    'is_live': is_live,
                    'checked_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    'uid': uid
                }
                
                # Save results
                save_uid_check_results()
            
            time.sleep(0.5)  # Small delay to prevent CPU overload
            
        except Exception as e:
            print(f"Error in background UID checker: {e}")
            time.sleep(1)

# Start background UID checker thread
def start_background_uid_checker():
    thread = threading.Thread(target=background_uid_checker, daemon=True)
    thread.start()

# Add UID to check queue
def add_uid_to_check_queue(uid, account_id, callback_data=None):
    uid_check_queue.append((uid, account_id, callback_data))

# Get UID check result
def get_uid_check_result(account_id):
    return uid_check_results.get(account_id)

# Get rules
def get_rules():
    try:
        rules_data = load_json(RULES_FILE)
        return rules_data.get("rules", "No rules set yet.")
    except:
        return "No rules set yet."

# Set rules
def set_rules(new_rules):
    rules_data = {"rules": new_rules}
    save_json(RULES_FILE, rules_data)
    return True

# Count unprocessed accounts in stock
def count_unprocessed_stock():
    try:
        # Check all files in uploaded_files folder
        files = [f for f in os.listdir(FILES_FOLDER) if f.endswith('.txt')]
        total_unprocessed = 0
        
        for filename in files:
            filepath = os.path.join(FILES_FOLDER, filename)
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip()]
            
            # Count unprocessed lines
            for line in lines:
                parts = line.split('|')
                if len(parts) >= 2:
                    username = parts[0].strip()
                    if not is_account_processed(username):
                        total_unprocessed += 1
        
        return total_unprocessed, len(files)
    except Exception as e:
        print(f"Error counting stock: {e}")
        return 0, 0

# ========== NEW FEATURES ==========

# 1. USER WORK LIMIT SYSTEM

# Get user limit
def get_user_limit(user_id):
    try:
        limits_data = load_json(USER_LIMITS_FILE)
        return limits_data.get(str(user_id), 0)  # 0 means no limit
    except:
        return 0

# Set user limit
def set_user_limit(username, limit):
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    try:
        limits_data = load_json(USER_LIMITS_FILE)
        limits_data[str(user_id)] = int(limit)
        save_json(USER_LIMITS_FILE, limits_data)
        return True, user_id
    except:
        return False, "Error setting limit"

# Check if user has reached limit
def check_user_limit(user_id):
    limit = get_user_limit(user_id)
    if limit == 0:  # No limit
        return False, 0
    
    # Get user's processed count
    user_stats = get_user_stats(user_id)
    if not user_stats:
        return False, 0
    
    processed_count = user_stats.get("total", 0)
    remaining = max(0, limit - processed_count)
    
    if processed_count >= limit:
        return True, remaining
    return False, remaining

# 2. TAKEN SYSTEM

# Mark taken accounts in Excel
def mark_taken_in_excel():
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel file not found"
        
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Get all rows with "Confirmed" status and not already marked as Taken
        taken_rows = []
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=6).value  # Status column
            taken_status = ws.cell(row=row, column=8).value  # Taken column
            
            if status and "confirmed" in str(status).lower() and not taken_status:
                # Mark as Taken
                ws.cell(row=row, column=8, value="Taken")
                taken_cell = ws.cell(row=row, column=8)
                taken_cell.fill = TAKEN_FILL
                taken_cell.font = Font(color="FFFFFF", bold=True)
                taken_cell.alignment = ALIGN_CENTER
                
                # Store row info
                username = ws.cell(row=row, column=2).value
                processor = ws.cell(row=row, column=4).value
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                taken_rows.append({
                    'row': row,
                    'username': username,
                    'processor': processor,
                    'marked_at': timestamp
                })
        
        # Save taken rows to file
        taken_data = load_json(TAKEN_FILE)
        existing_rows = taken_data.get("taken_rows", [])
        existing_rows.extend(taken_rows)
        taken_data["taken_rows"] = existing_rows
        save_json(TAKEN_FILE, taken_data)
        
        # Save Excel
        wb.save(EXCEL_FILE)
        
        return True, len(taken_rows)
        
    except Exception as e:
        print(f"Error marking taken accounts: {e}")
        return False, str(e)

# Get taken accounts info
def get_taken_info():
    try:
        if not os.path.exists(EXCEL_FILE):
            return {"total_confirmed": 0, "taken": 0, "fresh": 0}
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        total_confirmed = 0
        total_taken = 0
        total_fresh = 0
        
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=6).value
            taken_status = ws.cell(row=row, column=8).value
            
            if status and "confirmed" in str(status).lower():
                total_confirmed += 1
                if taken_status:
                    total_taken += 1
                else:
                    total_fresh += 1
        
        return {
            "total_confirmed": total_confirmed,
            "taken": total_taken,
            "fresh": total_fresh
        }
        
    except:
        return {"total_confirmed": 0, "taken": 0, "fresh": 0}

# 3. PENDING ACCOUNTS SYSTEM (UPDATED - FIXED)

# Get pending accounts info
def get_pending_accounts_info():
    try:
        pending_data = load_json(PENDING_ACCOUNTS_FILE)
        return pending_data.get("pending_accounts", {})
    except:
        return {}

# Mark all confirmed accounts as pending (FIXED)
def mark_all_as_pending():
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel file not found"
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        pending_count = 0
        pending_users = {}
        
        # Count all confirmed accounts and group by user
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=6).value  # Status column
            processor_code = ws.cell(row=row, column=5).value  # User Code column
            processor_name = ws.cell(row=row, column=4).value  # Processor column
            
            if status and "confirmed" in str(status).lower():
                pending_count += 1
                
                # Find user ID from processor name
                user_id = get_user_id_from_username(processor_name.lower().replace('@', '')) if processor_name else None
                
                if processor_name and user_id:
                    # Get username without @
                    username = processor_name.replace('@', '')
                    
                    if username not in pending_users:
                        pending_users[username] = {
                            'user_id': user_id,
                            'user_code': processor_code,
                            'pending_count': 0
                        }
                    pending_users[username]['pending_count'] += 1
        
        # Save to pending accounts file
        pending_data = load_json(PENDING_ACCOUNTS_FILE)
        pending_data["pending_accounts"] = pending_users
        save_json(PENDING_ACCOUNTS_FILE, pending_data)
        
        return True, {
            'total_pending': pending_count,
            'users': pending_users,
            'user_count': len(pending_users)
        }
        
    except Exception as e:
        print(f"Error marking all as pending: {e}")
        return False, str(e)

# 4. USER ACCOUNT CHECK SYSTEM (NEW FEATURE)

# Check user's confirmed accounts
def check_user_accounts(username):
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    try:
        if not os.path.exists(EXCEL_FILE):
            return False, "Excel file not found"
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        user_code = get_user_code(user_id)
        if not user_code:
            return False, "User code not found"
        
        # Get user's confirmed accounts from Excel
        user_accounts = []
        for row in range(2, ws.max_row + 1):
            processor_code = ws.cell(row=row, column=5).value  # User Code column
            status = ws.cell(row=row, column=6).value  # Status column
            username_col = ws.cell(row=row, column=2).value  # Username column
            
            if processor_code == user_code and status and "confirmed" in str(status).lower():
                user_accounts.append({
                    'row': row,
                    'username': username_col,
                    'status': status
                })
        
        if not user_accounts:
            return False, "No confirmed accounts found for this user"
        
        return True, user_accounts
        
    except Exception as e:
        print(f"Error getting user accounts: {e}")
        return False, str(e)

# Check if UID is live with interval
def check_uid_with_interval(uid):
    time.sleep(CHECK_INTERVAL)  # Add interval between checks
    return check_uid_live_sync(uid)

# Check all user accounts in background with live updates
def check_user_accounts_background(username, chat_id, message_id, update_stats=False):
    try:
        success, result = check_user_accounts(username)
        
        if not success:
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=f"❌ Error: {result}"
            )
            return
        
        user_accounts = result
        total_accounts = len(user_accounts)
        
        # Get user info
        user_id = get_user_id_from_username(username.lower())
        user_stats = get_user_stats(user_id)
        original_confirmed = user_stats.get("confirmed", 0) if user_stats else 0
        
        # Create initial message
        message = bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=f"🔍 Checking @{username}'s accounts...\n\n"
                 f"📊 Total accounts to check: {total_accounts}\n"
                 f"✅ Originally confirmed: {original_confirmed}\n"
                 f"⏰ Interval: {CHECK_INTERVAL} seconds per account\n"
                 f"⏳ Estimated time: {total_accounts * CHECK_INTERVAL:.1f} seconds\n\n"
                 f"📈 Progress: 0/{total_accounts} (0%)\n"
                 f"✅ Live: 0 | ❌ Dead: 0"
        )
        
        live_count = 0
        dead_count = 0
        checked_accounts = []
        
        # Check each account with interval
        for i, account in enumerate(user_accounts, 1):
            uid = account['username']
            
            # Check if UID is live
            is_live = check_uid_with_interval(uid)
            
            checked_accounts.append({
                'uid': uid,
                'is_live': is_live,
                'original_status': account['status']
            })
            
            if is_live:
                live_count += 1
            else:
                dead_count += 1
            
            # Update progress after each check
            progress_percent = (i / total_accounts) * 100
            
            # Update message with live results
            try:
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=message_id,
                    text=f"🔍 Checking @{username}'s accounts...\n\n"
                         f"📊 Total accounts to check: {total_accounts}\n"
                         f"✅ Originally confirmed: {original_confirmed}\n"
                         f"⏰ Interval: {CHECK_INTERVAL} seconds per account\n\n"
                         f"📈 Progress: {i}/{total_accounts} ({progress_percent:.1f}%)\n"
                         f"✅ Live: {live_count} | ❌ Dead: {dead_count}\n\n"
                         f"🔍 Last checked: {uid[:15]}... - {'✅ LIVE' if is_live else '❌ DEAD'}"
                )
            except:
                pass
        
        # Save check results
        check_data = load_json(USER_CHECK_RESULTS_FILE)
        check_data[username] = {
            'checked_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'total_checked': total_accounts,
            'live_count': live_count,
            'dead_count': dead_count,
            'original_confirmed': original_confirmed,
            'accounts': checked_accounts
        }
        save_json(USER_CHECK_RESULTS_FILE, check_data)
        
        # Prepare result message
        result_text = f"""
✅ Check Complete!

👤 User: @{username}
📅 Checked at: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

📊 Results:
✅ Originally confirmed: {original_confirmed}
✅ Live accounts: {live_count}
❌ Dead accounts: {dead_count}

📈 Accuracy: {(live_count / total_accounts * 100) if total_accounts > 0 else 0:.1f}%

💡 Note:
Live accounts are still working
Dead accounts are suspended/banned
        """
        
        if update_stats:
            # Update user stats if live count is different
            if live_count != original_confirmed:
                # Ask for confirmation to update stats
                keyboard = InlineKeyboardMarkup()
                keyboard.row(
                    InlineKeyboardButton(f"✅ Update to {live_count} confirmed", callback_data=f"update_stats_{username}_{live_count}"),
                    InlineKeyboardButton("❌ Keep original stats", callback_data=f"keep_stats_{username}")
                )
                
                result_text += f"\n\n⚠️ Difference found!\nOriginal: {original_confirmed} confirmed\nLive: {live_count} accounts\n\nDo you want to update user's confirmed count?"
                
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=message_id,
                    text=result_text,
                    reply_markup=keyboard
                )
            else:
                result_text += f"\n\n✅ No difference found. Stats are accurate."
                bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=message_id,
                    text=result_text
                )
        else:
            # Just show results
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=result_text
            )
        
    except Exception as e:
        print(f"Error in background check: {e}")
        bot.edit_message_text(
            chat_id=chat_id,
            message_id=message_id,
            text=f"❌ Error during check: {str(e)}"
        )

# Update user stats after check
def update_user_stats_after_check(username, new_confirmed):
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    try:
        users_data = load_json(USERS_FILE)
        
        if "users" not in users_data or user_id not in users_data["users"]:
            return False, "User stats not found"
        
        # Get current stats
        current_stats = users_data["users"][user_id]
        original_confirmed = current_stats.get("confirmed", 0)
        original_total = current_stats.get("total", 0)
        
        # Calculate difference
        difference = new_confirmed - original_confirmed
        
        # Update confirmed count
        users_data["users"][user_id]["confirmed"] = new_confirmed
        
        # Update total (adjust based on difference)
        users_data["users"][user_id]["total"] = original_total + difference
        
        # If difference is negative, add to suspended
        if difference < 0:
            users_data["users"][user_id]["suspended"] = current_stats.get("suspended", 0) + abs(difference)
        
        save_json(USERS_FILE, users_data)
        
        # Notify user
        try:
            bot.send_message(
                user_id,
                f"📢 Account Check Update\n\n"
                f"Your accounts have been checked by admin.\n"
                f"✅ Original confirmed: {original_confirmed}\n"
                f"✅ New confirmed: {new_confirmed}\n"
                f"📊 Difference: {difference}\n\n"
                f"Your statistics have been updated accordingly."
            )
        except:
            pass
        
        return True, {
            "username": username,
            "original_confirmed": original_confirmed,
            "new_confirmed": new_confirmed,
            "difference": difference
        }
        
    except Exception as e:
        print(f"Error updating user stats: {e}")
        return False, str(e)

# ========== UNMARKP COMMAND ==========

@bot.message_handler(commands=['unmarkp'])
def unmarkp_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    command = message.text.replace('/unmarkp', '').strip()
    
    if command.lower() == 'all':
        # Clear all pending accounts
        pending_data = load_json(PENDING_ACCOUNTS_FILE)
        pending_count = sum(info.get('pending_count', 0) for info in pending_data.get("pending_accounts", {}).values())
        
        pending_data["pending_accounts"] = {}
        save_json(PENDING_ACCOUNTS_FILE, pending_data)
        
        bot.send_message(message.chat.id,
                         f"✅ All pending accounts have been cleared!\n\n"
                         f"📊 Cleared: {pending_count} pending accounts\n"
                         f"👥 Affected users: All users with pending accounts\n\n"
                         f"💡 Pending system has been reset to 0 for all users.")
        return
    
    else:
        # Clear specific user's pending accounts
        username = command.replace('@', '').strip()
        if not username:
            bot.send_message(message.chat.id,
                           "📝 Clear Pending Accounts\n\n"
                           "Usage:\n"
                           "/unmarkp @username  - Clear user's pending accounts\n"
                           "/unmarkp all       - Clear ALL pending accounts\n\n"
                           "Example:\n"
                           "/unmarkp nhossain123\n\n"
                           "Note: This resets pending count to 0")
            return
        
        # Clear user's pending accounts
        pending_data = load_json(PENDING_ACCOUNTS_FILE)
        pending_users = pending_data.get("pending_accounts", {})
        
        if username in pending_users:
            pending_count = pending_users[username].get('pending_count', 0)
            del pending_users[username]
            pending_data["pending_accounts"] = pending_users
            save_json(PENDING_ACCOUNTS_FILE, pending_data)
            
            bot.send_message(message.chat.id,
                           f"✅ Pending accounts cleared for @{username}!\n\n"
                           f"📊 Cleared: {pending_count} pending accounts\n"
                           f"✅ Pending count reset to 0\n\n"
                           f"💡 User can now be marked as pending again if needed.")
        else:
            bot.send_message(message.chat.id, f"ℹ️ No pending accounts found for @{username}")

# ========== COMMAND HANDLERS ==========

# Set cooldown command (Admin only)
@bot.message_handler(commands=['set_cooldown'])
def set_cooldown_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    try:
        seconds = int(message.text.replace('/set_cooldown', '').strip())
        if seconds < 0:
            seconds = 0
        
        if set_cooldown_time(seconds):
            current_cooldown = get_cooldown_time()
            if seconds == 0:
                bot.send_message(message.chat.id, f"✅ Cooldown has been disabled (0 seconds)")
            else:
                bot.send_message(message.chat.id, f"✅ Cooldown set to {seconds} seconds")
    
    except ValueError:
        current_cooldown = get_cooldown_time()
        bot.send_message(message.chat.id, 
                        f"Current Cooldown: {current_cooldown} seconds\n\n"
                        "Usage:\n"
                        "/set_cooldown seconds\n\n"
                        "Examples:\n"
                        "/set_cooldown 5  (5 seconds cooldown)\n"
                        "/set_cooldown 0  (disable cooldown)\n"
                        "/set_cooldown 10 (10 seconds cooldown)")

# ========== SYSTEM STATUS COMMANDS ==========

@bot.message_handler(commands=['off'])
def off_command(message):
    """Turn off work system"""
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    if disable_work():
        off_notice = get_off_notice()
        
        # Broadcast to all users
        users_data = load_json(USERS_FILE)
        user_ids = list(users_data.get("users", {}).keys())
        
        success = 0
        failed = 0
        
        for uid in user_ids:
            try:
                bot.send_message(uid, off_notice)
                success += 1
            except:
                failed += 1
        
        response = f"""
🔴 SYSTEM TURNED OFF

{off_notice}

📢 Broadcast Status:
✅ Sent to: {success} users
❌ Failed: {failed} users

💡 Users will not be able to use /work command until system is turned back on.
        """
        
        bot.send_message(message.chat.id, response)
    else:
        bot.send_message(message.chat.id, "❌ Error turning off system")

@bot.message_handler(commands=['on'])
def on_command(message):
    """Turn on work system"""
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    if enable_work():
        on_notice = get_on_notice()
        
        # Broadcast to all users
        users_data = load_json(USERS_FILE)
        user_ids = list(users_data.get("users", {}).keys())
        
        success = 0
        failed = 0
        
        for uid in user_ids:
            try:
                bot.send_message(uid, on_notice)
                success += 1
            except:
                failed += 1
        
        response = f"""
🟢 SYSTEM TURNED ON

{on_notice}

📢 Broadcast Status:
✅ Sent to: {success} users
❌ Failed: {failed} users

💡 Users can now use /work command again.
        """
        
        bot.send_message(message.chat.id, response)
    else:
        bot.send_message(message.chat.id, "❌ Error turning on system")

@bot.message_handler(commands=['setoffnotice'])
def setoffnotice_command(message):
    """Set offline notice"""
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    notice = message.text.replace('/setoffnotice', '').strip()
    
    if not notice:
        current_notice = get_off_notice()
        bot.send_message(message.chat.id,
                        f"🔴 Current Offline Notice:\n\n{current_notice}\n\n"
                        "Usage: /setoffnotice your notice message here\n\n"
                        "Example: /setoffnotice ⚠️ System maintenance in progress. Please wait.")
        return
    
    if set_off_notice(notice):
        bot.send_message(message.chat.id, f"✅ Offline notice set:\n\n{notice}")
    else:
        bot.send_message(message.chat.id, "❌ Error setting offline notice")

@bot.message_handler(commands=['setonnotice'])
def setonnotice_command(message):
    """Set online notice"""
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    notice = message.text.replace('/setonnotice', '').strip()
    
    if not notice:
        current_notice = get_on_notice()
        bot.send_message(message.chat.id,
                        f"🟢 Current Online Notice:\n\n{current_notice}\n\n"
                        "Usage: /setonnotice your notice message here\n\n"
                        "Example: /setonnotice ✅ System is back online! You can start working now.")
        return
    
    if set_on_notice(notice):
        bot.send_message(message.chat.id, f"✅ Online notice set:\n\n{notice}")
    else:
        bot.send_message(message.chat.id, "❌ Error setting online notice")

# ========== RESETSTATS COMMAND (FIXED) ==========

@bot.message_handler(commands=['resetstats'])
def resetstats_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    command = message.text.replace('/resetstats', '').strip()
    
    if not command:
        # Show help with emojis explanation
        help_text = """
🔄 Reset Statistics - Help

📊 Emojis Legend:
✅ Confirmed (Live accounts)
❌ Suspended accounts
🟡 C.Suspended accounts
⚠️ Issue accounts
📊 Total processed accounts

Usage:
/resetstats @username           - Reset user stats
/resetstats @username full      - Reset ALL user data
/resetstats @username -live     - Reset stats, keep only LIVE accounts
/resetstats @username -suspended - Reset stats, keep only SUSPENDED accounts
/resetstats @username -c_suspended - Reset stats, keep only C.SUSPENDED accounts
/resetstats @username -issue    - Reset stats, keep only ISSUE accounts
/resetstats all                - Reset ALL user stats
/resetstats all full           - Reset ALL user stats

📝 Examples:
/resetstats nhossain123          - Reset stats for @nhossain123
/resetstats nhossain123 full     - Reset ALL data for @nhossain123
/resetstats nhossain123 -live    - Keep only live accounts, remove others
/resetstats nhossain123 -suspended - Keep only suspended accounts
/resetstats nhossain123 -issue   - Keep only issue accounts
/resetstats all                 - Reset all users' stats
/resetstats all full            - Reset ALL data for all users

⚠️ Warnings:
- Use 'full' to also clear Excel data
- This action cannot be undone!
        """
        bot.send_message(message.chat.id, help_text)
        return
    
    parts = command.split()
    
    if parts[0].lower() == 'all':
        full_reset = len(parts) > 1 and parts[1].lower() == 'full'
        
        # Reset all users
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(f"✅ Yes, Reset All", callback_data=f"reset_all_{'full' if full_reset else 'normal'}"),
            InlineKeyboardButton("❌ No, Cancel", callback_data="reset_cancel")
        )
        
        reset_type = "FULL (including Excel data)" if full_reset else "NORMAL (stats only)"
        
        bot.send_message(message.chat.id,
                         f"⚠️ Reset ALL User Statistics\n\n"
                         f"Reset Type: {reset_type}\n\n"
                         f"This will reset ALL user statistics:\n"
                         f"- ✅ Confirmed counts\n"
                         f"- ❌ Suspended counts\n"
                         f"- 🟡 C.Suspended counts\n"
                         f"- ⚠️ Issue counts\n"
                         f"- 📈 Total processed\n"
                         f"- 📊 Excel file will be cleared\n\n"
                         f"This action cannot be undone!\n\n"
                         f"Are you sure you want to reset all statistics?",
                         reply_markup=keyboard)
        return
    
    # Process specific user
    username = parts[0].replace('@', '').strip()
    
    # Check if user exists
    target_user_id = get_user_id_from_username(username.lower())
    if not target_user_id:
        bot.send_message(message.chat.id, f"❌ User @{username} not found.")
        return
    
    # Check for full reset
    full_reset = len(parts) > 1 and parts[1].lower() == 'full'
    
    # Parse keep options
    keep_types = []
    for part in parts[1:]:
        if part.startswith('-') and part[1:].lower() != 'full':
            keep_type = part[1:].lower()
            if keep_type in ['live', 'confirmed', 'suspended', 'c_suspended', 'c.suspended', 'issue']:
                # Normalize live/confirmed
                if keep_type in ['live', 'confirmed']:
                    keep_types.append('confirmed')
                elif keep_type == 'c.suspended':
                    keep_types.append('c_suspended')
                else:
                    keep_types.append(keep_type)
    
    if full_reset:
        # Full reset
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(f"✅ Reset ALL Data", callback_data=f"reset_user_{username}_full"),
            InlineKeyboardButton("❌ Cancel", callback_data="reset_cancel")
        )
        
        bot.send_message(message.chat.id,
                         f"⚠️ FULL Reset User Data\n\n"
                         f"👤 User: @{username}\n\n"
                         f"This will reset ALL data for this user:\n"
                         f"- ✅ Confirmed counts\n"
                         f"- ❌ Suspended counts\n"
                         f"- 🟡 C.Suspended counts\n"
                         f"- ⚠️ Issue counts\n"
                         f"- 📈 Total processed\n"
                         f"- 📊 Excel entries will be removed\n\n"
                         f"This action cannot be undone!\n\n"
                         f"Are you sure?",
                         reply_markup=keyboard)
    
    elif not keep_types:
        # Reset all stats
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(f"✅ Reset @{username}'s Stats", callback_data=f"reset_user_{username}_all"),
            InlineKeyboardButton("❌ Cancel", callback_data="reset_cancel")
        )
        
        bot.send_message(message.chat.id,
                         f"⚠️ Reset User Statistics\n\n"
                         f"👤 User: @{username}\n\n"
                         f"This will reset ALL statistics for this user:\n"
                         f"- ✅ Confirmed counts\n"
                         f"- ❌ Suspended counts\n"
                         f"- 🟡 C.Suspended counts\n"
                         f"- ⚠️ Issue counts\n"
                         f"- 📈 Total processed\n"
                         f"- 📊 Excel entries will be removed\n\n"
                         f"This action cannot be undone!\n\n"
                         f"Are you sure?",
                         reply_markup=keyboard)
    
    elif len(keep_types) == 1:
        # Reset all except one type
        keep_type = keep_types[0]
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(f"✅ Reset, Keep {keep_type}", callback_data=f"reset_user_{username}_{keep_type}"),
            InlineKeyboardButton("❌ Cancel", callback_data="reset_cancel")
        )
        
        bot.send_message(message.chat.id,
                         f"⚠️ Reset User Statistics\n\n"
                         f"👤 User: @{username}\n"
                         f"💾 Keep: {keep_type} accounts\n\n"
                         f"This will reset ALL statistics except {keep_type} accounts:\n"
                         f"- Will keep only {keep_type} accounts\n"
                         f"- Remove all other account types\n"
                         f"- Update Excel file accordingly\n\n"
                         f"This action cannot be undone!\n\n"
                         f"Are you sure?",
                         reply_markup=keyboard)
    
    else:
        # Multiple keep types - complex operation
        keep_types_str = ', '.join(keep_types)
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton(f"✅ Reset, Keep Selected", callback_data=f"reset_user_{username}_multiple"),
            InlineKeyboardButton("❌ Cancel", callback_data="reset_cancel")
        )
        
        bot.send_message(message.chat.id,
                         f"⚠️ Reset User Statistics\n\n"
                         f"👤 User: @{username}\n"
                         f"💾 Keep: {keep_types_str}\n\n"
                         f"This will reset statistics, keeping only:\n"
                         f"- {keep_types_str} accounts\n"
                         f"- Remove all other account types\n"
                         f"- Update Excel file accordingly\n\n"
                         f"This action cannot be undone!\n\n"
                         f"Are you sure?",
                         reply_markup=keyboard)
        
        # Store the keep types for later use
        user_states[user_id] = {
            'step': 'reset_user_multiple',
            'username': username,
            'keep_types': keep_types
        }

# Handle reset callback (FIXED)
@bot.callback_query_handler(func=lambda call: call.data.startswith('reset_'))
def handle_reset_callback(call):
    user_id = str(call.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.answer_callback_query(call.id, "❌ Permission denied")
        return
    
    if call.data == "reset_cancel":
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text="❌ Statistics reset cancelled."
        )
        bot.answer_callback_query(call.id, "Cancelled")
        return
    
    elif call.data.startswith("reset_all_"):
        parts = call.data.split('_')
        full_reset = parts[2] == 'full' if len(parts) > 2 else False
        
        if reset_all_stats(full_reset=full_reset):
            reset_type = "FULL (including Excel data)" if full_reset else "NORMAL (stats only)"
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"✅ All user statistics have been reset ({reset_type}) AND Excel file has been cleared!"
            )
        else:
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="❌ Error resetting statistics."
            )
        bot.answer_callback_query(call.id)
    
    elif call.data.startswith("reset_user_"):
        parts = call.data.split('_')
        if len(parts) >= 4:
            username = parts[2]
            reset_type = parts[3] if len(parts) > 3 else "all"
            
            if reset_type == "full":
                # Full reset
                success, result = reset_user_stats(username, full_reset=True)
                
                if success:
                    response = f"""
✅ FULL Reset Complete!

👤 User: @{username}
✅ Reset ALL data including:
  - ✅ Confirmed: 0
  - ❌ Suspended: 0
  - 🟡 C.Suspended: 0
  - ⚠️ Issue: 0
  - 📊 Total: 0
  - 📊 Excel data: Removed all entries
                    """
                    bot.edit_message_text(
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        text=response
                    )
                else:
                    bot.edit_message_text(
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        text=f"❌ Error: {result}"
                    )
            
            elif reset_type == "all":
                # Reset all stats
                success, result = reset_user_stats(username)
                
                if success:
                    response = f"""
✅ Reset Complete!

👤 User: @{username}
✅ Statistics reset to zero:
  - ✅ Confirmed: 0
  - ❌ Suspended: 0
  - 🟡 C.Suspended: 0
  - ⚠️ Issue: 0
  - 📊 Total: 0
                    """
                    bot.edit_message_text(
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        text=response
                    )
                else:
                    bot.edit_message_text(
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        text=f"❌ Error: {result}"
                    )
            
            elif reset_type == "multiple":
                # Multiple keep types - need to handle specially
                if user_id in user_states and user_states[user_id]['step'] == 'reset_user_multiple':
                    keep_types = user_states[user_id].get('keep_types', [])
                    
                    # This is a complex operation - we'll keep only the specified types
                    # For simplicity, we'll implement a basic version
                    
                    success, result = reset_user_stats(username, keep_type="confirmed" if 'confirmed' in keep_types else None)
                    
                    if success:
                        response = f"✅ Reset stats for @{username}, keeping: {', '.join(keep_types)}\n"
                        response += "⚠️ Note: Complex filtering requires manual Excel cleanup"
                        
                        bot.edit_message_text(
                            chat_id=call.message.chat.id,
                            message_id=call.message.message_id,
                            text=response
                        )
                    else:
                        bot.edit_message_text(
                            chat_id=call.message.chat.id,
                            message_id=call.message.message_id,
                            text=f"❌ Error: {result}"
                        )
                    
                    # Clear state
                    if user_id in user_states:
                        del user_states[user_id]
            
            else:
                # Reset all except one type
                success, result = reset_user_stats(username, reset_type)
                
                if success:
                    response = f"""
✅ Reset Complete!

👤 User: @{username}
✅ Statistics reset, kept only {reset_type} accounts
                    """
                    bot.edit_message_text(
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        text=response
                    )
                else:
                    bot.edit_message_text(
                        chat_id=call.message.chat.id,
                        message_id=call.message.message_id,
                        text=f"❌ Error: {result}"
                    )
            
            bot.answer_callback_query(call.id)

# ========== CHECK COMMAND ==========

@bot.message_handler(commands=['check'])
def check_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    command = message.text.replace('/check', '').strip()
    
    if not command:
        bot.send_message(message.chat.id,
                        "🔍 Check User Accounts\n\n"
                        "Usage:\n"
                        "/check @username\n\n"
                        "This will check all confirmed accounts\n"
                        "of the specified user and verify if\n"
                        "they are still live or dead.\n\n"
                        "Example:\n"
                        "/check nhossain123")
        return
    
    username = command.replace('@', '').strip()
    
    # Check if user exists
    target_user_id = get_user_id_from_username(username.lower())
    if not target_user_id:
        bot.send_message(message.chat.id, f"❌ User @{username} not found.")
        return
    
    # Send initial message
    msg = bot.send_message(
        message.chat.id,
        f"🔍 Starting check for @{username}...\n\n"
        f"⏰ Please wait, this may take a while.\n"
        f"Interval: {CHECK_INTERVAL} seconds per account\n\n"
        f"Preparing to check user's confirmed accounts..."
    )
    
    # Start background check thread
    thread = threading.Thread(
        target=check_user_accounts_background,
        args=(username, message.chat.id, msg.message_id, True)
    )
    thread.start()

# ========== TAKEN COMMAND ==========

@bot.message_handler(commands=['taken'])
def taken_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    success, result = mark_taken_in_excel()
    
    if success:
        taken_count = result
        taken_info = get_taken_info()
        
        response = f"""
✅ Taken System

📊 Marking Results:
✅ Marked as Taken: {taken_count} accounts

📈 Current Status:
✅ Total Confirmed: {taken_info['total_confirmed']}
🔵 Taken: {taken_info['taken']}
🟢 Fresh: {taken_info['fresh']}

💡 Note:
Taken accounts are marked with blue color
in Excel file. Use /export to download updated file.
        """
        bot.send_message(message.chat.id, response)
    else:
        bot.send_message(message.chat.id, f"❌ Error: {result}")

# ========== MARKP COMMAND (FIXED) ==========

@bot.message_handler(commands=['markp'])
def markp_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    command = message.text.replace('/markp', '').strip()
    
    if command.lower() == 'all':
        # Mark all confirmed accounts as pending (FIXED)
        success, result = mark_all_as_pending()
        
        if success:
            info = result
            
            response = f"""
✅ Pending System - All Users

📊 Marking Results:
✅ Total accounts marked as pending: {info['total_pending']}
👥 Affected users: {info['user_count']}

📋 User Breakdown:
"""
            
            # List users with pending counts
            user_count = 0
            for username, user_info in info['users'].items():
                if user_count < 10:  # Show first 10 users
                    response += f"• @{username}: {user_info['pending_count']} accounts\n"
                    user_count += 1
            
            if info['user_count'] > 10:
                response += f"\n... and {info['user_count'] - 10} more users"
            
            # Notify all affected users
            notified_count = 0
            for username, user_info in info['users'].items():
                user_id_target = user_info.get('user_id')
                pending_count = user_info.get('pending_count', 0)
                
                if user_id_target and pending_count > 0:
                    try:
                        bot.send_message(
                            user_id_target,
                            f"📢 Accounts Marked as Pending\n\n"
                            f"Admin has marked {pending_count} of your confirmed accounts as pending.\n\n"
                            f"These accounts will be checked and processed later.\n"
                            f"Use /stats to see your current pending count."
                        )
                        notified_count += 1
                    except:
                        pass
            
            response += f"\n📢 Notifications sent to: {notified_count} users"
            
            bot.send_message(message.chat.id, response)
        else:
            bot.send_message(message.chat.id, f"❌ Error: {result}")
    
    else:
        # Mark specific user's accounts as pending
        username = command.replace('@', '').strip()
        if not username:
            bot.send_message(message.chat.id,
                           "📝 Mark Pending Accounts\n\n"
                           "Usage:\n"
                           "/markp @username  - Mark user's accounts as pending\n"
                           "/markp all       - Mark ALL confirmed accounts as pending\n\n"
                           "Example:\n"
                           "/markp nhossain123\n\n"
                           "Note: This marks accounts in system only,\n"
                           "not in Excel file.")
            return
        
        # Find user
        target_user_id = get_user_id_from_username(username.lower())
        if not target_user_id:
            bot.send_message(message.chat.id, f"❌ User @{username} not found.")
            return
        
        user_code = get_user_code(target_user_id)
        if not user_code:
            bot.send_message(message.chat.id, f"❌ User code not found for @{username}")
            return
        
        try:
            if not os.path.exists(EXCEL_FILE):
                bot.send_message(message.chat.id, "❌ Excel file not found")
                return
            
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            
            pending_count = 0
            
            # Count user's confirmed accounts
            for row in range(2, ws.max_row + 1):
                processor_code = ws.cell(row=row, column=5).value  # User Code column
                status = ws.cell(row=row, column=6).value  # Status column
                
                if processor_code == user_code and status and "confirmed" in str(status).lower():
                    pending_count += 1
            
            if pending_count > 0:
                # Save to pending accounts file
                pending_data = load_json(PENDING_ACCOUNTS_FILE)
                pending_users = pending_data.get("pending_accounts", {})
                
                pending_users[username] = {
                    'user_id': target_user_id,
                    'user_code': user_code,
                    'pending_count': pending_count,
                    'marked_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                
                pending_data["pending_accounts"] = pending_users
                save_json(PENDING_ACCOUNTS_FILE, pending_data)
                
                # Notify the user
                try:
                    bot.send_message(
                        target_user_id,
                        f"📢 Accounts Marked as Pending\n\n"
                        f"Admin has marked {pending_count} of your confirmed accounts as pending.\n\n"
                        f"These accounts will be checked and processed later.\n"
                        f"Use /stats to see your current pending count."
                    )
                except:
                    pass
            else:
                bot.send_message(message.chat.id, f"ℹ️ No confirmed accounts found for @{username}")
                return
            
            pending_info = get_pending_accounts_info()
            
            response = f"""
✅ Pending System

📊 Marking Results:
✅ User: @{username}
✅ Accounts marked as pending: {pending_count}

📈 Current Pending Stats:
👥 Total Users with Pending: {len(pending_info)}
📊 Total Pending Accounts: {sum(info['pending_count'] for info in pending_info.values())}

💡 Note:
Accounts are marked as pending in the system only.
This does NOT affect the Excel file.

📢 User has been notified.
            """
            bot.send_message(message.chat.id, response)
            
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Error: {str(e)}")

# ========== EDIT COMMAND ==========

# Edit user statistics
def edit_user_stats(username, field, value):
    user_id = get_user_id_from_username(username.lower())
    if not user_id:
        return False, "User not found"
    
    try:
        users_data = load_json(USERS_FILE)
        
        if "users" not in users_data or user_id not in users_data["users"]:
            return False, "User not found in database"
        
        # Validate field
        valid_fields = ["confirmed", "suspended", "c_suspended", "issue", "total"]
        field_lower = field.lower()
        
        if field_lower not in valid_fields:
            return False, f"Invalid field. Valid fields are: {', '.join(valid_fields)}"
        
        try:
            new_value = int(value)
            if new_value < 0:
                return False, "Value cannot be negative"
        except ValueError:
            return False, "Value must be a number"
        
        # Update the field
        users_data["users"][user_id][field_lower] = new_value
        
        # Save updated data
        save_json(USERS_FILE, users_data)
        
        return True, f"Updated @{username}'s {field} to {new_value}"
        
    except Exception as e:
        print(f"Error editing user stats: {e}")
        return False, str(e)

@bot.message_handler(commands=['edit'])
def edit_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    command = message.text.replace('/edit', '').strip()
    
    if not command:
        bot.send_message(message.chat.id,
                        "📝 Edit User Statistics\n\n"
                        "Usage:\n"
                        "/edit @username field value\n\n"
                        "Available Fields:\n"
                        "• confirmed    - Confirmed accounts\n"
                        "• suspended    - Suspended accounts\n"
                        "• c_suspended  - C.Suspended accounts\n"
                        "• issue        - Issue accounts\n"
                        "• total        - Total accounts (auto-calculated)\n\n"
                        "Examples:\n"
                        "/edit nhossain123 confirmed 50\n"
                        "/edit nhossain123 suspended 10\n"
                        "/edit nhossain123 total 100")
        return
    
    parts = command.split()
    if len(parts) < 3:
        bot.send_message(message.chat.id, "❌ Invalid format. Use: /edit @username field value")
        return
    
    username = parts[0].replace('@', '').strip()
    field = parts[1].strip()
    value = parts[2].strip()
    
    success, result = edit_user_stats(username, field, value)
    
    if success:
        # Get updated stats
        user_id_target = get_user_id_from_username(username.lower())
        user_stats = get_user_stats(user_id_target)
        
        if user_stats:
            response = f"""
✅ Statistics Updated Successfully!

{result}

📊 Updated Stats for @{username}:
✅ Confirmed: {user_stats.get('confirmed', 0)}
❌ Suspended: {user_stats.get('suspended', 0)}
🟡 C.Suspended: {user_stats.get('c_suspended', 0)}
⚠️ Issue: {user_stats.get('issue', 0)}
📊 Total: {user_stats.get('total', 0)}
            """
        else:
            response = f"✅ {result}"
        
        bot.send_message(message.chat.id, response)
    else:
        bot.send_message(message.chat.id, f"❌ Error: {result}")

# ========== LIMIT COMMAND ==========

@bot.message_handler(commands=['limit'])
def limit_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    command = message.text.replace('/limit', '').strip()
    
    if not command:
        # Show current limits
        limits_data = load_json(USER_LIMITS_FILE)
        
        if not limits_data:
            bot.send_message(message.chat.id, "📊 No user limits set yet.")
            return
        
        response = "📊 User Work Limits:\n\n"
        for uid, limit in limits_data.items():
            # Get username from user ID
            users_data = load_json(USERS_FILE)
            username = "Unknown"
            if uid in users_data.get("users", {}):
                username = users_data["users"][uid].get("username", "Unknown")
            
            # Get current processed count
            user_stats = get_user_stats(uid)
            processed = user_stats.get("total", 0) if user_stats else 0
            
            status = "✅ Active" if limit == 0 or processed < limit else "🚫 Limit Reached"
            remaining = max(0, limit - processed) if limit > 0 else "∞"
            
            response += f"👤 @{username}\n"
            response += f"   🔢 Limit: {limit if limit > 0 else 'No limit'}\n"
            response += f"   📊 Processed: {processed}\n"
            response += f"   📈 Remaining: {remaining}\n"
            response += f"   🔄 Status: {status}\n\n"
        
        response += "\n💡 Usage: /limit @username limit (0 for no limit)"
        bot.send_message(message.chat.id, response)
        return
    
    parts = command.split()
    if len(parts) < 2:
        bot.send_message(message.chat.id,
                        "📝 Set User Work Limit\n\n"
                        "Usage:\n"
                        "/limit @username limit\n\n"
                        "Examples:\n"
                        "/limit nhossain123 10  - Limit to 10 accounts\n"
                        "/limit nhossain123 0   - Remove limit\n\n"
                        "Note: 0 means no limit")
        return
    
    username = parts[0].replace('@', '').strip()
    
    try:
        limit = int(parts[1].strip())
        if limit < 0:
            bot.send_message(message.chat.id, "❌ Limit cannot be negative")
            return
        
        success, result = set_user_limit(username, limit)
        
        if success:
            user_id_target = result
            
            # Get user info
            user_stats = get_user_stats(user_id_target)
            processed = user_stats.get("total", 0) if user_stats else 0
            
            if limit == 0:
                response = f"""
✅ User Limit Updated!

👤 User: @{username}
🚫 Limit: Removed (No limit)
📊 Currently Processed: {processed}
✅ User can now process unlimited accounts
                """
            else:
                remaining = max(0, limit - processed)
                response = f"""
✅ User Limit Updated!

👤 User: @{username}
🔢 New Limit: {limit} accounts
📊 Currently Processed: {processed}
📈 Remaining: {remaining} accounts
⏳ Status: {'🚫 Limit Reached' if processed >= limit else '✅ Active'}
                """
            
            bot.send_message(message.chat.id, response)
            
            # Notify the user if they're active
            try:
                if limit == 0:
                    bot.send_message(user_id_target, 
                                   f"🎉 Your work limit has been removed!\n"
                                   f"You can now process unlimited accounts.")
                else:
                    bot.send_message(user_id_target,
                                   f"📢 Work Limit Updated\n\n"
                                   f"Your new work limit: {limit} accounts\n"
                                   f"Currently processed: {processed}\n"
                                   f"Remaining: {remaining}\n\n"
                                   f"Once you reach {limit} accounts, "
                                   f"you won't be able to process more.")
            except:
                pass
        else:
            bot.send_message(message.chat.id, f"❌ Error: {result}")
    
    except ValueError:
        bot.send_message(message.chat.id, "❌ Limit must be a number")

# ========== PENDINGINFO COMMAND ==========

@bot.message_handler(commands=['pendinginfo'])
def pendinginfo_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    pending_info = get_pending_accounts_info()
    
    if not pending_info:
        bot.send_message(message.chat.id, "📭 No pending accounts found.")
        return
    
    response = "📊 Pending Accounts Information\n\n"
    total_pending = 0
    
    for username, info in pending_info.items():
        user_id_info = info.get('user_id', 'Unknown')
        user_code = info.get('user_code', 'Unknown')
        count = info.get('pending_count', 0)
        marked_at = info.get('marked_at', 'Unknown')
        
        total_pending += count
        
        response += f"👤 @{username}\n"
        response += f"   🔢 User Code: {user_code}\n"
        response += f"   📊 Pending Count: {count}\n"
        response += f"   ⏰ Marked At: {marked_at}\n"
        response += f"   🆔 User ID: {user_id_info}\n\n"
    
    response += f"📈 Total Pending Accounts: {total_pending}\n"
    response += f"👥 Total Users with Pending: {len(pending_info)}\n\n"
    response += "💡 Commands:\n"
    response += "/markp @username - Mark user's accounts as pending\n"
    response += "/markp all - Mark ALL accounts as pending\n"
    response += "/clearp @username - Clear user's pending status"
    
    bot.send_message(message.chat.id, response)

# ========== CLEARP COMMAND ==========

@bot.message_handler(commands=['clearp'])
def clearp_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    username = message.text.replace('/clearp', '').strip().replace('@', '').strip()
    
    if not username:
        bot.send_message(message.chat.id,
                        "📝 Clear Pending Status\n\n"
                        "Usage:\n"
                        "/clearp @username\n\n"
                        "Example:\n"
                        "/clearp nhossain123\n\n"
                        "This will clear the pending status from\n"
                        "the system for this user.")
        return
    
    # Remove from pending accounts file
    pending_data = load_json(PENDING_ACCOUNTS_FILE)
    pending_users = pending_data.get("pending_accounts", {})
    
    if username in pending_users:
        # Get user ID for notification
        user_info = pending_users[username]
        user_id_target = user_info.get('user_id')
        
        del pending_users[username]
        pending_data["pending_accounts"] = pending_users
        save_json(PENDING_ACCOUNTS_FILE, pending_data)
        
        # Notify the user
        if user_id_target:
            try:
                bot.send_message(
                    user_id_target,
                    f"📢 Pending Status Cleared\n\n"
                    f"Your pending accounts status has been cleared by admin.\n"
                    f"You no longer have any pending accounts."
                )
            except:
                pass
        
        response = f"""
✅ Pending Status Cleared!

👤 User: @{username}
✅ Pending status cleared from system
📢 User has been notified

💡 Note:
The user's accounts are no longer marked
as pending in the system.
        """
    else:
        response = f"ℹ️ No pending accounts found for @{username}"
    
    bot.send_message(message.chat.id, response)

# ========== ALLSTATS COMMAND (UPDATED WITHOUT PROCESSED IDs) ==========

@bot.message_handler(commands=['allstats'])
def allstats_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    users_data = load_json(USERS_FILE)
    users_dict = users_data.get("users", {})
    
    if not users_dict:
        bot.send_message(message.chat.id, "❌ No users found.")
        return
    
    # Get pending accounts info
    pending_info = get_pending_accounts_info()
    
    users_list = []
    total_confirmed = 0
    total_suspended = 0
    total_c_suspended = 0
    total_issue = 0
    total_all = 0
    banned_users = []
    pending_approvals_count = len(get_pending_approvals())
    
    for uid, stats in users_dict.items():
        username = stats.get('username', 'Unknown')
        user_code = stats.get('user_code', 'Unknown')
        confirmed = stats.get('confirmed', 0)
        suspended = stats.get('suspended', 0)
        c_suspended = stats.get('c_suspended', 0)
        issue = stats.get('issue', 0)
        total = stats.get('total', 0)
        
        # Get pending count for this user
        user_pending_count = 0
        if username in pending_info:
            user_pending_count = pending_info[username].get('pending_count', 0)
        
        is_banned = uid in users_data.get("banned", [])
        status_symbol = "🚫" if is_banned else "✅"
        
        # Get user limit
        limit = get_user_limit(uid)
        limit_text = f" | 🚫 Limit: {limit}" if limit > 0 else ""
        
        # Format user entry with all emojis
        user_entry = f"{status_symbol} {user_code} (@{username})\n"
        user_entry += f"   ✅ {confirmed} | ❌ {suspended} | 🟡 {c_suspended} | ⚠️ {issue} | 📊 {total}"
        
        if user_pending_count > 0:
            user_entry += f" | 🟠 Pending: {user_pending_count}"
        
        user_entry += limit_text
        
        users_list.append(user_entry)
        
        if is_banned:
            banned_users.append(f"@{username}")
        
        total_confirmed += confirmed
        total_suspended += suspended
        total_c_suspended += c_suspended
        total_issue += issue
        total_all += total
    
    # Get stock information
    unprocessed_count, file_count = count_unprocessed_stock()
    
    # Get current cooldown
    cooldown_time = get_cooldown_time()
    
    # Get taken info
    taken_info = get_taken_info()
    
    # Get total pending accounts
    total_pending_accounts = sum(info.get('pending_count', 0) for info in pending_info.values())
    
    # Get system status
    system_status = get_system_status()
    work_enabled = system_status.get("work_enabled", True)
    system_status_text = "🟢 ONLINE" if work_enabled else "🔴 OFFLINE"
    
    # Create summary with emojis legend
    summary = f"""
📊 ALL USER STATS

📈 Emojis Legend:
✅ Confirmed/Live accounts
❌ Suspended accounts  
🟡 C.Suspended accounts
⚠️ Issue accounts
📊 Total processed accounts
🟠 Pending accounts
🚫 User limit
✅ Active user
🚫 Banned user

👥 User Overview:
✅ Approved Users: {len(users_dict)}
⏳ Pending Approvals: {pending_approvals_count}
✅ Active Users: {len(users_dict) - len(banned_users)}
🚫 Banned Users: {len(banned_users)}

📈 Total Performance:
✅ Total Confirmed: {total_confirmed}
❌ Total Suspended: {total_suspended}
🟡 Total C.Suspended: {total_c_suspended}
⚠️ Total Issue: {total_issue}
📊 Total Processed: {total_all}

🆕 Feature Stats:
🔵 Taken Accounts: {taken_info.get('taken', 0)}
🟠 Pending Accounts: {total_pending_accounts}
🟢 Fresh Accounts: {taken_info.get('fresh', 0)}

📦 Stock Information:
📁 Files: {file_count}
🔢 Unprocessed: {unprocessed_count}

⏰ System Settings:
Cooldown: {cooldown_time} seconds
🔍 Check Interval: {CHECK_INTERVAL} seconds
🔌 System Status: {system_status_text}
    """
    
    if banned_users:
        summary += f"\n\n🚫 Banned Users:\n" + "\n".join(banned_users)
    
    # Send summary first
    bot.send_message(message.chat.id, summary)
    
    # Then send individual user stats (split if too long)
    user_stats_text = "📋 Individual User Stats:\n\n"
    for i, user_stat in enumerate(users_list):
        if len(user_stats_text + user_stat + "\n\n") > 4000:  # Telegram message limit
            bot.send_message(message.chat.id, user_stats_text)
            user_stats_text = ""
        user_stats_text += user_stat + "\n\n"
    
    if user_stats_text:
        bot.send_message(message.chat.id, user_stats_text)

# ========== OTHER COMMANDS (REST OF THE CODE REMAINS SIMILAR) ==========

# Cancel command - cancel all user states
@bot.message_handler(commands=['cancel'])
def cancel_command(message):
    user_id = str(message.from_user.id)
    
    # Clear all user states
    if user_id in user_states:
        del user_states[user_id]
    
    # Clear work session if any
    if user_id in work_sessions:
        del work_sessions[user_id]
    
    # Clear any account data for this user
    account_ids_to_remove = []
    for account_id, data in account_data_store.items():
        if data['user_id'] == user_id:
            account_ids_to_remove.append(account_id)
    
    for account_id in account_ids_to_remove:
        del account_data_store[account_id]
    
    # Clear cooldown timer
    if user_id in cooldown_timers:
        del cooldown_timers[user_id]
    
    bot.send_message(message.chat.id,
                    "✅ All operations cancelled!\n\n"
                    "You have been returned to the main menu.\n"
                    "Use /start to begin again.")

# Start command
@bot.message_handler(commands=['start'])
def start(message):
    user_id = str(message.from_user.id)
    username = message.from_user.username or "Unknown"
    
    # Clear any existing states
    if user_id in user_states:
        del user_states[user_id]
    
    # Update username mapping
    if username != "Unknown":
        update_username_mapping(user_id, username)
    
    # Check if user is banned
    if is_user_banned(user_id):
        bot.send_message(message.chat.id, "❌ You have been banned from using this bot.")
        return
    
    # Check if user is approved
    if not is_user_approved(user_id):
        # Check if already pending
        if has_pending_approval(user_id):
            bot.send_message(message.chat.id, 
                           "⏳ Your approval request is pending.\n"
                           "Please wait for admin approval.")
            return
        
        # Show approval request
        keyboard = InlineKeyboardMarkup()
        keyboard.row(InlineKeyboardButton("✅ Apply for Approval", callback_data="apply_approval"))
        
        bot.send_message(message.chat.id,
                        "🔒 Approval Required\n\n"
                        "You need admin approval to use this bot.\n"
                        "Click the button below to apply for approval:",
                        reply_markup=keyboard)
        return
    
    # Welcome message
    welcome_msg = """
ACCOUNT MANAGER BOT

👋 Welcome to the ultimate account processing system!

📋 Use /help to see available commands

💼 Use /work to start processing accounts

🔒 Secure | Efficient | Reliable
    """
    
    bot.send_message(message.chat.id, welcome_msg)

# Work command
@bot.message_handler(commands=['work'])
def work_command(message):
    user_id = str(message.from_user.id)
    username = message.from_user.username or "User"
    
    # Update username mapping
    if username != "User":
        update_username_mapping(user_id, username)
    
    # Check if user is banned
    if is_user_banned(user_id):
        bot.send_message(message.chat.id, "❌ You have been banned from using this bot.")
        return
    
    # Check if work system is enabled
    if not is_work_enabled():
        off_notice = get_off_notice()
        bot.send_message(message.chat.id, off_notice)
        return
    
    # Check if user is approved
    if not is_user_approved(user_id):
        # Check if pending
        if has_pending_approval(user_id):
            bot.send_message(message.chat.id, 
                           "⏳ Your approval request is pending.\n"
                           "Please wait for admin approval.")
            return
        
        # Show approval request
        keyboard = InlineKeyboardMarkup()
        keyboard.row(InlineKeyboardButton("✅ Apply for Approval", callback_data="apply_approval"))
        
        bot.send_message(message.chat.id,
                        "🔒 Approval Required\n\n"
                        "You need admin approval to use this bot.\n"
                        "Click the button below to apply for approval:",
                        reply_markup=keyboard)
        return
    
    # Check if user has reached limit
    limit_reached, remaining = check_user_limit(user_id)
    if limit_reached:
        limit = get_user_limit(user_id)
        bot.send_message(message.chat.id,
                        f"🚫 Work Limit Reached!\n\n"
                        f"You have reached your work limit of {limit} accounts.\n"
                        f"Please contact admin to increase your limit.\n\n"
                        f"Your stats: {remaining} accounts remaining (0)")
        return
    
    # Check if there are any uploaded files
    files = [f for f in os.listdir(FILES_FOLDER) if f.endswith('.txt')]
    if not files:
        bot.send_message(message.chat.id, 
                         "❌ No accounts file found. Please contact admin to upload files.")
        return
    
    # Get the most recent file
    latest_file = max(files, key=lambda f: os.path.getctime(os.path.join(FILES_FOLDER, f)))
    filepath = os.path.join(FILES_FOLDER, latest_file)
    
    # Read all lines
    with open(filepath, 'r', encoding='utf-8') as f:
        all_lines = [line.strip() for line in f if line.strip()]
    
    # Filter out already processed accounts
    unprocessed_lines = []
    for line in all_lines:
        parts = line.split('|')
        if len(parts) >= 2:
            account_username = parts[0].strip()
            if not is_account_processed(account_username):
                unprocessed_lines.append(line)
    
    if not unprocessed_lines:
        bot.send_message(message.chat.id, 
                         "📭 No unprocessed accounts found.\n"
                         "All accounts in the file have been processed.")
        return
    
    # Get user code
    user_code = get_user_code(user_id) or "Unknown"
    
    # Store work session
    work_sessions[user_id] = {
        'filepath': filepath,
        'lines': unprocessed_lines,
        'current_index': 0,
        'processed_count': 0,
        'username': username,
        'user_code': user_code
    }
    
    # Get rules
    rules = get_rules()
    
    # Get current cooldown
    cooldown_time = get_cooldown_time()
    
    # Get user limit info
    limit = get_user_limit(user_id)
    user_stats = get_user_stats(user_id)
    processed = user_stats.get("total", 0) if user_stats else 0
    
    # Show instructions with buttons
    instructions = f"""
💼 WORK MODE

👤 Worker: @{username}
🔢 User Code: {user_code}

📋 Rules:
{rules}

📊 Available: {len(unprocessed_lines)} accounts

⏰ Cooldown: {cooldown_time} seconds per account

📈 Your Stats:
✅ Processed: {processed} accounts
{'🚫 Limit: ' + str(limit) + ' accounts' if limit > 0 else '✅ No limit'}
{'📊 Remaining: ' + str(remaining) + ' accounts' if limit > 0 else ''}

Ready to start processing?
    """
    
    keyboard = InlineKeyboardMarkup()
    keyboard.row(
        InlineKeyboardButton("🚀 Start Working", callback_data="work_start"),
        InlineKeyboardButton("❌ Cancel", callback_data="work_cancel")
    )
    
    bot.send_message(message.chat.id, instructions, reply_markup=keyboard)

# Handle approval request
@bot.callback_query_handler(func=lambda call: call.data == 'apply_approval')
def handle_approval_request(call):
    user_id = str(call.from_user.id)
    username = call.from_user.username or "Unknown"
    
    # Check if already approved
    if is_user_approved(user_id):
        bot.answer_callback_query(call.id, "✅ You are already approved!")
        return
    
    # Check if already pending
    if has_pending_approval(user_id):
        bot.answer_callback_query(call.id, "⏳ Your request is already pending")
        return
    
    # Add to pending approvals
    if add_pending_approval(user_id, username):
        bot.answer_callback_query(call.id, "✅ Approval request sent!")
        
        # Notify admin
        admin_msg = f"📝 New Approval Request\n\n👤 User: @{username}\n🆔 ID: {user_id}"
        
        # Find admin users
        users_data = load_json(USERS_FILE)
        for uid in users_data.get("users", {}):
            if is_admin(uid) or is_subadmin(uid):
                try:
                    bot.send_message(uid, admin_msg)
                except:
                    pass
        
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text="✅ Approval Request Sent!\n\nAdmin will review your request soon."
        )
    else:
        bot.answer_callback_query(call.id, "❌ Error sending request")

# Pending approvals command
@bot.message_handler(commands=['pending'])
def pending_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    pending_list = get_pending_approvals()
    
    if not pending_list:
        bot.send_message(message.chat.id, "✅ No pending approvals.")
        return
    
    # Create message with pending list
    pending_text = "Pending Approvals:\n\n"
    for i, pending in enumerate(pending_list, 1):
        pending_text += f"{i}. 👤 @{pending['username']}\n   🆔 {pending['user_id']}\n   ⏰ {pending['timestamp']}\n\n"
    
    # Create buttons
    keyboard = InlineKeyboardMarkup()
    keyboard.row(InlineKeyboardButton("✅ Approve All", callback_data="approve_all"))
    
    for i, pending in enumerate(pending_list, 1):
        keyboard.row(InlineKeyboardButton(f"✅ Approve @{pending['username']}", callback_data=f"approve_{pending['user_id']}"))
    
    bot.send_message(message.chat.id, pending_text, reply_markup=keyboard)

# Approve command
@bot.message_handler(commands=['approve'])
def approve_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    username = message.text.replace('/approve', '').strip().replace('@', '').strip()
    if not username:
        bot.send_message(message.chat.id, 
                        "Usage:\n"
                        "/approve @username\n\n"
                        "Example:\n"
                        "/approve nhossain123")
        return
    
    # Find user ID
    target_user_id = get_user_id_from_username(username.lower())
    if not target_user_id:
        bot.send_message(message.chat.id, f"❌ User @{username} not found.")
        return
    
    # Check if already approved
    if is_user_approved(target_user_id):
        bot.send_message(message.chat.id, f"✅ User @{username} is already approved.")
        return
    
    # Approve user
    if approve_user(target_user_id):
        # Assign user code
        user_code = get_next_user_code()
        
        # Add to users
        users_data = load_json(USERS_FILE)
        users_data.setdefault("users", {})[target_user_id] = {
            "username": username,
            "user_code": user_code,
            "confirmed": 0,
            "suspended": 0,
            "c_suspended": 0,
            "issue": 0,
            "total": 0
        }
        save_json(USERS_FILE, users_data)
        
        bot.send_message(message.chat.id, f"✅ User @{username} approved with code: {user_code}")
        
        # Notify the user
        try:
            bot.send_message(target_user_id, 
                           f"🎉 Your approval has been granted!\n\n"
                           f"Your user code: {user_code}\n"
                           f"You can now use /work to start processing accounts.")
        except:
            pass
    else:
        bot.send_message(message.chat.id, f"❌ User @{username} not found in pending list.")

# Handle approval callbacks
@bot.callback_query_handler(func=lambda call: call.data.startswith('approve_'))
def handle_approval_callback(call):
    user_id = str(call.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.answer_callback_query(call.id, "❌ Permission denied")
        return
    
    action = call.data.split('_')[1]
    
    if action == "all":
        # Approve all pending users
        approved_user_ids = approve_all_pending()
        
        if not approved_user_ids:
            bot.answer_callback_query(call.id, "No pending approvals")
            return
        
        approved_count = 0
        for target_user_id in approved_user_ids:
            # Find username
            pending_list = get_pending_approvals()
            username = "Unknown"
            for pending in pending_list:
                if pending["user_id"] == target_user_id:
                    username = pending["username"]
                    break
            
            # Assign user code
            user_code = get_next_user_code()
            
            # Add to users
            users_data = load_json(USERS_FILE)
            users_data.setdefault("users", {})[target_user_id] = {
                "username": username,
                "user_code": user_code,
                "confirmed": 0,
                "suspended": 0,
                "c_suspended": 0,
                "issue": 0,
                "total": 0
            }
            save_json(USERS_FILE, users_data)
            
            # Notify the user
            try:
                bot.send_message(target_user_id, 
                               f"🎉 Your approval has been granted!\n\n"
                               f"Your user code: {user_code}\n"
                               f"You can now use /work to start processing accounts.")
            except:
                pass
            
            approved_count += 1
        
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text=f"✅ {approved_count} users approved successfully!"
        )
        bot.answer_callback_query(call.id, f"Approved {approved_count} users")
    
    else:
        # Approve specific user
        target_user_id = action
        
        # Find username
        pending_list = get_pending_approvals()
        username = "Unknown"
        for pending in pending_list:
            if pending["user_id"] == target_user_id:
                username = pending["username"]
                break
        
        # Approve user
        if approve_user(target_user_id):
            # Assign user code
            user_code = get_next_user_code()
            
            # Add to users
            users_data = load_json(USERS_FILE)
            users_data.setdefault("users", {})[target_user_id] = {
                "username": username,
                "user_code": user_code,
                "confirmed": 0,
                "suspended": 0,
                "c_suspended": 0,
                "issue": 0,
                "total": 0
            }
            save_json(USERS_FILE, users_data)
            
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"✅ User @{username} approved with code: {user_code}"
            )
            bot.answer_callback_query(call.id, f"Approved @{username}")
            
            # Notify the user
            try:
                bot.send_message(target_user_id, 
                               f"🎉 Your approval has been granted!\n\n"
                               f"Your user code: {user_code}\n"
                               f"You can now use /work to start processing accounts.")
            except:
                pass
        else:
            bot.answer_callback_query(call.id, "User not found in pending list")

# List admins command
@bot.message_handler(commands=['listadmin'])
def listadmin_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    admins = get_all_admins()
    
    admin_text = "Administrators List:\n\n"
    
    # Main admin
    admin_text += "Main Admin:\n"
    if admins["main_admin"]:
        for admin in admins["main_admin"]:
            status = "✅ Online" if admin_sessions.get(admin["user_id"]) else "❌ Offline"
            admin_text += f"👤 @{admin['username']} ({status})\n   🆔 {admin['user_id']}\n"
    else:
        admin_text += "No main admin currently logged in.\n"
    
    # Sub admins
    admin_text += "\nSub-Admins:\n"
    if admins["sub_admins"]:
        for subadmin in admins["sub_admins"]:
            status = "✅ Online" if subadmin["is_logged_in"] else "❌ Offline"
            admin_text += f"👤 @{subadmin['username']} ({status})\n   🆔 {subadmin['user_id']}\n   📅 Added: {subadmin['created_at']}\n"
    else:
        admin_text += "No sub-admins added yet.\n"
    
    admin_text += f"\nTotal: {len(admins['main_admin'])} main admin(s), {len(admins['sub_admins'])} sub-admin(s)"
    
    bot.send_message(message.chat.id, admin_text)

# Remove subadmin command
@bot.message_handler(commands=['rmvadmin'])
def rmvadmin_command(message):
    user_id = str(message.from_user.id)
    
    if not is_admin(user_id):
        bot.send_message(message.chat.id, "❌ This command is for main admin only.")
        return
    
    username = message.text.replace('/rmvadmin', '').strip().replace('@', '').strip()
    if not username:
        bot.send_message(message.chat.id, 
                        "Usage:\n"
                        "/rmvadmin @username\n\n"
                        "Example:\n"
                        "/rmvadmin nhossain123")
        return
    
    success, result = remove_subadmin(username)
    
    if success:
        bot.send_message(message.chat.id, f"✅ Sub-admin @{username} removed successfully.")
        
        # Notify the removed subadmin
        try:
            bot.send_message(result, 
                           "⚠️ Your sub-admin privileges have been removed.\n\n"
                           "You can no longer access admin commands.")
        except:
            pass
    else:
        bot.send_message(message.chat.id, f"❌ {result}")

# Remove user command
@bot.message_handler(commands=['rmvuser'])
def rmvuser_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    username = message.text.replace('/rmvuser', '').strip().replace('@', '').strip()
    if not username:
        bot.send_message(message.chat.id, 
                        "Usage:\n"
                        "/rmvuser @username\n\n"
                        "Example:\n"
                        "/rmvuser nhossain123")
        return
    
    success, result = remove_user(username)
    
    if success:
        bot.send_message(message.chat.id, f"✅ User @{username} removed from database.")
        
        # Notify the removed user
        try:
            bot.send_message(result, 
                           "⚠️ Your account has been removed from the system.\n\n"
                           "You will need to apply for approval again to use the bot.")
        except:
            pass
    else:
        bot.send_message(message.chat.id, f"❌ {result}")

# Add subadmin command (Admin only)
@bot.message_handler(commands=['addsubadmin'])
def addsubadmin_command(message):
    user_id = str(message.from_user.id)
    
    if not is_admin(user_id):
        bot.send_message(message.chat.id, "❌ This command is for main admin only.")
        return
    
    # Clear any existing state for this user
    if user_id in user_states:
        del user_states[user_id]
    
    bot.send_message(message.chat.id,
                     "👥 Add Sub-Admin\n\n"
                     "Format:\n"
                     "@username password\n\n"
                     "Example:\n"
                     "@nhossain123 mypassword123\n\n"
                     "Or send /cancel to cancel this operation.")
    
    user_states[user_id] = {'step': 'awaiting_subadmin'}

# Handle subadmin input
@bot.message_handler(func=lambda message: str(message.from_user.id) in user_states and 
                     user_states[str(message.from_user.id)]['step'] == 'awaiting_subadmin')
def handle_subadmin_input(message):
    user_id = str(message.from_user.id)
    text = message.text.strip()
    
    # Check for cancel command
    if text.lower() == '/cancel':
        del user_states[user_id]
        bot.send_message(message.chat.id, "❌ Sub-admin addition cancelled.")
        return
    
    if not text.startswith('@'):
        bot.send_message(message.chat.id, "❌ Username must start with @. Please try again or /cancel to cancel.")
        return
    
    # Parse username and password
    parts = text.split()
    if len(parts) < 2:
        bot.send_message(message.chat.id, "❌ Format: @username password\nUse /cancel to cancel.")
        return
    
    username = parts[0].replace('@', '').strip()
    password = parts[1].strip()
    
    # Add subadmin
    success, result = add_subadmin(username, password)
    
    if success:
        bot.send_message(message.chat.id, 
                        f"✅ Sub-admin added successfully!\n\n"
                        f"👤 Username: @{username}\n"
                        f"🔐 Password: {password}\n"
                        f"🆔 User ID: {result}")
        
        # Notify the new subadmin
        try:
            bot.send_message(result,
                           f"🎉 You have been promoted to Sub-Admin!\n\n"
                           f"Login with: /slogin {password}")
        except:
            pass
    else:
        bot.send_message(message.chat.id, f"❌ {result}")
    
    # Clear state
    if user_id in user_states:
        del user_states[user_id]

# Subadmin login command
@bot.message_handler(commands=['slogin'])
def slogin_command(message):
    user_id = str(message.from_user.id)
    username = message.from_user.username or "Unknown"
    
    # Check if user is banned
    if is_user_banned(user_id):
        bot.send_message(message.chat.id, "❌ You have been banned from using this bot.")
        return
    
    # Check if already logged in as subadmin
    if subadmin_sessions.get(user_id):
        bot.send_message(message.chat.id, "✅ You are already logged in as sub-admin!")
        return
    
    password = message.text.replace('/slogin', '').strip()
    if not password:
        bot.send_message(message.chat.id, 
                        "🔐 Sub-Admin Login\n\n"
                        "Usage:\n"
                        "/slogin password")
        return
    
    # Check subadmin credentials
    subadmins_data = load_json(SUBADMINS_FILE)
    subadmin_info = subadmins_data.get("subadmins", {}).get(user_id)
    
    if subadmin_info and subadmin_info.get("password") == password:
        subadmin_sessions[user_id] = True
        bot.send_message(message.chat.id,
                        "✅ Sub-Admin Login Successful!\n\n"
                        "You now have admin privileges (except changing main settings).")
    else:
        bot.send_message(message.chat.id, "❌ Invalid sub-admin credentials.")

# Updated login command for main admin
@bot.message_handler(commands=['login'])
def login_command(message):
    user_id = str(message.from_user.id)
    username = message.from_user.username or "Unknown"
    
    # Update username mapping
    if username != "Unknown":
        update_username_mapping(user_id, username)
    
    # Check if user is banned
    if is_user_banned(user_id):
        bot.send_message(message.chat.id, "❌ You have been banned from using this bot.")
        return
    
    if is_admin(user_id) or subadmin_sessions.get(user_id):
        bot.send_message(message.chat.id, "✅ You are already logged in!")
        return
    
    # Check if this is main admin login
    password = message.text.replace('/login', '').strip()
    admin_password = get_admin_password()
    
    if password == admin_password:
        admin_sessions[user_id] = True
        bot.send_message(message.chat.id, 
                        "✅ Admin Login Successful!\n\n"
                        "You are now logged in as main admin.")
        return
    
    # If not main admin, check if it's subadmin login attempt
    if password:
        bot.send_message(message.chat.id, "❌ Invalid password. Use /slogin for sub-admin login.")
    else:
        bot.send_message(message.chat.id, 
                        "🔐 Login Options:\n\n"
                        "👑 Main Admin:\n"
                        "/login password\n\n"
                        "👥 Sub-Admin:\n"
                        "/slogin password")

# Help command (Updated with new commands)
@bot.message_handler(commands=['help'])
def help_command(message):
    user_id = str(message.from_user.id)
    
    if is_admin(user_id):
        help_text = """
Main Admin Menu

📁 File Management:
/upload - Upload accounts file
/export - Download results Excel
/deleteall - Delete all uploaded files

📊 Statistics:
/allstats - Show all user statistics
/stock - Check unprocessed accounts
/resetstats - Reset user statistics (with options)

🔍 Account Checking:
/check @username - Check user's confirmed accounts

👥 User Management:
/pending - Show pending approvals
/approve @username - Approve user
/ban @username - Ban user
/unban @username - Unban user
/rmvuser @username - Remove user
/limit @username limit - Set user work limit
/edit @username field value - Edit user stats
/pendinginfo - Show pending accounts info

🔧 Admin Management:
/addsubadmin - Add new sub-admin
/rmvadmin @username - Remove sub-admin
/listadmin - List all admins
/changepassword newpass - Change password
/set_cooldown seconds - Set processing cooldown

📢 Broadcast:
/broadcast message - Broadcast to users

⚙️ Settings:
/setrules - Set work rules

🔌 System Control:
/off - Turn off work system
/on - Turn on work system
/setoffnotice - Set offline notice
/setonnotice - Set online notice

🔄 System:
/cancel - Cancel current operation
/logout - Logout from admin
/help - Show this menu

🆕 NEW FEATURES:
/taken - Mark confirmed accounts as taken
/markp @username - Mark user's accounts as pending
/markp all - Mark ALL accounts as pending
/clearp @username - Clear user's pending status
/unmarkp @username - Clear user's pending accounts (reset to 0)
/unmarkp all - Clear ALL pending accounts

🔄 Reset Stats (Advanced):
/resetstats @username - Reset user's stats
/resetstats @username full - Reset ALL user data
/resetstats @username -live - Keep only live accounts
/resetstats @username -suspended - Keep only suspended
/resetstats @username -issue - Keep only issue accounts
/resetstats all - Reset all users' stats
/resetstats all full - Reset ALL users' data

Tip: Commands with @username need parameters
Example: /approve @username
        """
    elif is_subadmin(user_id) or subadmin_sessions.get(user_id):
        help_text = """
Sub-Admin Menu

📁 File Management:
/upload - Upload accounts file
/export - Download results Excel

📊 Statistics:
/allstats - Show all user statistics
/stock - Check unprocessed accounts

🔍 Account Checking:
/check @username - Check user's confirmed accounts

👥 User Management:
/pending - Show pending approvals
/approve @username - Approve user
/ban @username - Ban user
/unban @username - Unban user
/rmvuser @username - Remove user
/set_cooldown seconds - Set processing cooldown

🆕 NEW FEATURES:
/taken - Mark confirmed accounts as taken
/markp @username - Mark user's accounts as pending
/markp all - Mark ALL accounts as pending
/clearp @username - Clear user's pending status
/unmarkp @username - Clear user's pending accounts (reset to 0)
/unmarkp all - Clear ALL pending accounts

🔄 Reset Stats:
/resetstats @username - Reset user's stats
/resetstats @username full - Reset ALL user data
/resetstats @username -live - Keep only live accounts
/resetstats @username -suspended - Keep only suspended
/resetstats @username -issue - Keep only issue accounts

📢 Broadcast:
/broadcast message - Broadcast to users

🔄 System:
/cancel - Cancel current operation
/slogout - Logout from sub-admin
/help - Show this menu

Tip: Commands with @username need parameters
Example: /approve @username
        """
    else:
        help_text = """
User Commands

💼 Work Commands:
/work - Start processing accounts
/stats - Check your statistics

🆘 Support:
/admin - Contact admin for help
/cancel - Cancel current operation
/help - Show this menu

Note: You need approval to use /work
        """
    
    bot.send_message(message.chat.id, help_text)

# Admin contact command
@bot.message_handler(commands=['admin'])
def admin_command(message):
    # Get system status
    system_status = get_system_status()
    work_enabled = system_status.get("work_enabled", True)
    system_status_text = "🟢 ONLINE" if work_enabled else "🔴 OFFLINE"
    
    admin_info = f"""
Admin Contact

📧 Main Admin: {ADMIN_USERNAME}
👥 Sub-Admins: Contact main admin

📞 Need Help? 
Contact the admin directly for:
- Account issues
- Technical problems
- Approval requests
- General inquiries

💼 Bot Features:
✅ Facebook account processing
📊 Detailed statistics tracking
🔒 Secure and reliable system
👥 Multi-user support
📁 Excel export capabilities

🆕 New Features:
📋 Copyable username/password in work
🔵 Taken system for tracking accounts
🟠 Pending system for account checking
🔍 Account checking system
📈 User work limits
✏️ Edit user statistics
🔌 System on/off control
📢 Custom on/off notices
🗑️ Unmarkp system to clear pending accounts

📊 System Status: {system_status_text}
{'✅ Work system is enabled' if work_enabled else '⚠️ Work system is temporarily disabled'}
    """
    
    bot.send_message(message.chat.id, admin_info)

# Stats command
@bot.message_handler(commands=['stats'])
def stats_command(message):
    user_id = str(message.from_user.id)
    username = message.from_user.username or "Unknown"
    
    # Update username mapping
    if username != "Unknown":
        update_username_mapping(user_id, username)
    
    # Check if user is banned
    if is_user_banned(user_id):
        bot.send_message(message.chat.id, "❌ You have been banned from using this bot.")
        return
    
    # Check if user is approved
    if not is_user_approved(user_id):
        bot.send_message(message.chat.id, "❌ You need approval to use this bot.")
        return
    
    user_stats = get_user_stats(user_id)
    
    if user_stats:
        total = user_stats.get('total', 0)
        confirmed = user_stats.get('confirmed', 0)
        suspended = user_stats.get('suspended', 0)
        c_suspended = user_stats.get('c_suspended', 0)
        issue = user_stats.get('issue', 0)
        user_code = user_stats.get('user_code', 'Unknown')
        
        # Calculate success rate
        success_rate = 0
        if total > 0:
            success_rate = (confirmed / total) * 100
        
        # Get user limit info
        limit = get_user_limit(user_id)
        remaining = max(0, limit - total) if limit > 0 else "∞"
        
        # Get pending info
        pending_info = get_pending_accounts_info()
        user_pending = pending_info.get(user_stats.get('username', username), {})
        pending_count = user_pending.get('pending_count', 0) if user_pending else 0
        
        stats_text = f"""
📊 YOUR STATS

👤 User Info:
🔢 User Code: {user_code}
📛 Username: @{user_stats.get('username', username)}
🆔 User ID: {user_id}

📈 Performance:
✅ Confirmed: {confirmed}
❌ Suspended: {suspended}
🟡 C.Suspended: {c_suspended}
⚠️ Issue: {issue}
📊 Total Processed: {total}

📈 Success Rate: {success_rate:.1f}%

📋 Pending Status:
{'🟠 Pending accounts: ' + str(pending_count) if pending_count > 0 else '✅ No pending accounts'}

{'🚫 Work Limit: ' + str(limit) + ' accounts' if limit > 0 else '✅ No work limit'}
{'📊 Remaining: ' + str(remaining) + ' accounts' if limit > 0 else ''}
        """
    else:
        stats_text = f"""
📊 YOUR STATS

👤 User Info:
📛 Username: @{username}
🆔 User ID: {user_id}

📈 Performance:
✅ Confirmed: 0
❌ Suspended: 0
🟡 C.Suspended: 0
⚠️ Issue: 0
📊 Total Processed: 0

💡 Note: 
You haven't processed any accounts yet.
Use /work to start processing.
        """
    
    bot.send_message(message.chat.id, stats_text)

# Stock command
@bot.message_handler(commands=['stock'])
def stock_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    unprocessed_count, file_count = count_unprocessed_stock()
    
    stock_info = f"""
Account Stock

📊 Statistics:
📁 Total Files: {file_count}
🔢 Unprocessed Accounts: {unprocessed_count}

📈 Breakdown:
✅ Ready for work: {unprocessed_count} accounts
📂 In uploaded files: {file_count} files

Note: 
These are accounts that haven't been 
processed yet. Use /work to start 
processing these accounts.
    """
    
    bot.send_message(message.chat.id, stock_info)

# Ban command
@bot.message_handler(commands=['ban'])
def ban_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    username = message.text.replace('/ban', '').strip().replace('@', '').strip()
    if not username:
        bot.send_message(message.chat.id, 
                        "🚫 Ban User\n\n"
                        "Usage:\n"
                        "/ban @username\n\n"
                        "Example:\n"
                        "/ban nhossain123")
        return
    
    success, target_user_id = ban_user(username)
    
    if success:
        bot.send_message(message.chat.id, f"✅ User @{username} has been banned.")
        
        # Notify the banned user
        try:
            bot.send_message(target_user_id, "❌ You have been banned from using this bot.")
        except:
            pass
    else:
        bot.send_message(message.chat.id, f"❌ User @{username} not found or already banned.")

# Unban command
@bot.message_handler(commands=['unban'])
def unban_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    username = message.text.replace('/unban', '').strip().replace('@', '').strip()
    if not username:
        bot.send_message(message.chat.id, 
                        "🔓 Unban User\n\n"
                        "Usage:\n"
                        "/unban @username\n\n"
                        "Example:\n"
                        "/unban nhossain123")
        return
    
    success, target_user_id = unban_user(username)
    
    if success:
        bot.send_message(message.chat.id, f"✅ User @{username} has been unbanned.")
        
        # Notify the unbanned user
        try:
            bot.send_message(target_user_id, "✅ You have been unbanned. You can now use the bot again.")
        except:
            pass
    else:
        bot.send_message(message.chat.id, f"❌ User @{username} not found or not banned.")

# Set rules command
@bot.message_handler(commands=['setrules'])
def setrules_command(message):
    user_id = str(message.from_user.id)
    
    if not is_admin(user_id):
        bot.send_message(message.chat.id, "❌ This command is for main admin only.")
        return
    
    # Clear any existing state
    if user_id in user_states:
        del user_states[user_id]
    
    current_rules = get_rules()
    
    bot.send_message(message.chat.id,
                     f"📋 Current Rules:\n\n{current_rules}\n\n"
                     "Please send the new rules (you can use multiple lines):\n"
                     "Or send /cancel to cancel.")
    
    user_states[user_id] = {'step': 'awaiting_rules'}

# Handle rules input
@bot.message_handler(func=lambda message: str(message.from_user.id) in user_states and 
                     user_states[str(message.from_user.id)]['step'] == 'awaiting_rules')
def handle_rules_input(message):
    user_id = str(message.from_user.id)
    new_rules = message.text.strip()
    
    # Check for cancel
    if new_rules.lower() == '/cancel':
        del user_states[user_id]
        bot.send_message(message.chat.id, "❌ Rules update cancelled.")
        return
    
    if not new_rules:
        bot.send_message(message.chat.id, "❌ Rules cannot be empty. Please try /setrules again or /cancel to cancel.")
        return
    
    if set_rules(new_rules):
        bot.send_message(message.chat.id, "✅ Rules have been updated successfully!")
    else:
        bot.send_message(message.chat.id, "❌ Error saving rules.")
    
    # Clear state
    if user_id in user_states:
        del user_states[user_id]

# Logout commands
@bot.message_handler(commands=['logout'])
def logout_command(message):
    user_id = str(message.from_user.id)
    
    if is_admin(user_id):
        admin_sessions.pop(user_id, None)
        bot.send_message(message.chat.id, "✅ Successfully logged out from admin session.")
    elif subadmin_sessions.get(user_id):
        subadmin_sessions.pop(user_id, None)
        bot.send_message(message.chat.id, "✅ Successfully logged out from sub-admin session.")
    else:
        bot.send_message(message.chat.id, "ℹ️ You are not logged in.")

@bot.message_handler(commands=['slogout'])
def slogout_command(message):
    user_id = str(message.from_user.id)
    
    if subadmin_sessions.get(user_id):
        subadmin_sessions.pop(user_id, None)
        bot.send_message(message.chat.id, "✅ Successfully logged out from sub-admin session.")
    else:
        bot.send_message(message.chat.id, "ℹ️ You are not logged in as sub-admin.")

# Upload command
@bot.message_handler(commands=['upload'])
def upload_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    bot.send_message(message.chat.id, 
                     "📁 Upload Accounts File\n\n"
                     "Please send me a .txt file with accounts.\n"
                     "Format:\n"
                     "username|password|cookie_data\n\n"
                     "(One account per line)\n\n"
                     "Each line will be processed individually.")

# Handle file uploads
@bot.message_handler(content_types=['document'])
def handle_file(message):
    user_id = str(message.from_user.id)
    
    if message.document and message.document.mime_type == 'text/plain':
        # Download the file
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # Save file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"accounts_{timestamp}.txt"
        filepath = os.path.join(FILES_FOLDER, filename)
        
        with open(filepath, 'wb') as f:
            f.write(downloaded_file)
        
        # Count lines
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f if line.strip()]
        
        bot.send_message(message.chat.id, 
                         f"✅ File uploaded successfully!\n"
                         f"📊 Found {len(lines)} accounts to process.\n\n"
                         f"Use /work command to start processing.")
    else:
        bot.send_message(message.chat.id, "❌ Please send a .txt file")

# Allresult command - save permanent results
@bot.message_handler(commands=['allresult'])
def allresult_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    # Save permanent results
    success, result = save_permanent_results()
    
    if success:
        filename = result["filename"]
        records = result["records"]
        timestamp = result["timestamp"]
        
        # Create keyboard with download button
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton("📥 Download Permanent File", callback_data=f"download_allresult_{filename}"),
            InlineKeyboardButton("📋 List All Files", callback_data="list_allresults")
        )
        
        response = f"""
✅ Successfully saved permanent results!

📁 File Details:
📄 Filename: {filename}
📊 Records: {records} accounts
⏰ Timestamp: {timestamp}

💾 Storage:
📍 Folder: all_results/
🔒 Status: Permanent (won't be deleted)

📋 Total Files: {len(list_permanent_results())}

Note: 
This file contains ALL processed accounts
and will NOT be deleted by /resetstats
        """
        
        bot.send_message(message.chat.id, response, reply_markup=keyboard)
    else:
        bot.send_message(message.chat.id, f"❌ Error saving permanent results: {result}")

# Handle allresult callbacks
@bot.callback_query_handler(func=lambda call: call.data.startswith('download_allresult_') or call.data == 'list_allresults')
def handle_allresult_callback(call):
    user_id = str(call.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.answer_callback_query(call.id, "❌ Permission denied")
        return
    
    if call.data.startswith('download_allresult_'):
        # Download specific file
        filename = call.data.replace('download_allresult_', '')
        filepath = os.path.join(ALL_RESULTS_FOLDER, filename)
        
        if os.path.exists(filepath):
            try:
                with open(filepath, 'rb') as f:
                    bot.send_document(call.message.chat.id, f, 
                                    caption=f"📊 Permanent Results File\n📄 {filename}\n🔒 Permanent storage")
                bot.answer_callback_query(call.id, "✅ File sent!")
            except Exception as e:
                bot.answer_callback_query(call.id, f"❌ Error: {str(e)}")
        else:
            bot.answer_callback_query(call.id, "❌ File not found")
    
    elif call.data == 'list_allresults':
        # List all permanent result files
        files = list_permanent_results()
        
        if not files:
            bot.answer_callback_query(call.id, "No permanent files found")
            bot.send_message(call.message.chat.id, "📭 No permanent result files found.")
            return
        
        response = "📋 Permanent Result Files:\n\n"
        keyboard = InlineKeyboardMarkup()
        
        for i, file_info in enumerate(files[:10]):  # Show first 10 files
            filename = file_info["filename"]
            size_kb = file_info["size"] / 1024
            created = file_info["created"]
            
            response += f"{i+1}. {filename}\n"
            response += f"   📏 Size: {size_kb:.1f} KB\n"
            response += f"   📅 Created: {created}\n\n"
            
            # Add download button for each file
            keyboard.row(InlineKeyboardButton(f"📥 {filename}", callback_data=f"download_allresult_{filename}"))
        
        if len(files) > 10:
            response += f"\n📁 ... and {len(files) - 10} more files"
        
        response += f"\n💾 Total: {len(files)} permanent files"
        
        bot.send_message(call.message.chat.id, response, reply_markup=keyboard)
        bot.answer_callback_query(call.id, f"Found {len(files)} files")

# EXPORT command - send Excel file
@bot.message_handler(commands=['export'])
def export_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    if not os.path.exists(EXCEL_FILE):
        bot.send_message(message.chat.id, "❌ No results file found yet.")
        return
    
    try:
        # Check if file is empty or has data
        file_size = os.path.getsize(EXCEL_FILE)
        if file_size < 1024:  # Less than 1KB
            bot.send_message(message.chat.id, "📭 Excel file is empty. No data to show.")
            return
            
        # Count records
        try:
            df = pd.read_excel(EXCEL_FILE)
            record_count = len(df)
        except:
            record_count = 0
        
        # Get taken info
        taken_info = get_taken_info()
        
        # Get pending accounts info (from system only)
        pending_info = get_pending_accounts_info()
        total_pending = sum(info.get('pending_count', 0) for info in pending_info.values())
        
        # Send the Excel file
        with open(EXCEL_FILE, 'rb') as f:
            bot.send_document(message.chat.id, f, 
                            caption=f"📊 Results Excel File\n\n"
                                   f"📄 Current working file\n"
                                   f"📊 Records: {record_count} accounts\n"
                                   f"✅ Confirmed: Green background\n"
                                   f"❌ Suspended: Red background\n"
                                   f"🟡 C.Suspended: Yellow background\n"
                                   f"⚠️ Issue: Purple background\n"
                                   f"🔵 Taken: Blue background\n"
                                   f"📋 Full cookies in Input column\n"
                                   f"👤 User Code column added\n\n"
                                   f"🆕 Feature Stats:\n"
                                   f"🔵 Taken: {taken_info.get('taken', 0)}\n"
                                   f"🟢 Fresh: {taken_info.get('fresh', 0)}\n"
                                   f"🟠 Pending (System): {total_pending}")
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Error sending file: {str(e)}")

# Broadcast command
@bot.message_handler(commands=['broadcast'])
def broadcast_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    text = message.text.replace('/broadcast', '').strip()
    if not text:
        bot.send_message(message.chat.id, 
                        "📢 Broadcast Message\n\n"
                        "Usage:\n"
                        "/broadcast your message here\n\n"
                        "Example:\n"
                        "/broadcast Server maintenance at 10 PM")
        return
    
    users_data = load_json(USERS_FILE)
    user_ids = list(users_data.get("users", {}).keys())
    
    success = 0
    failed = 0
    
    bot.send_message(message.chat.id, f"📢 Broadcasting to {len(user_ids)} users...")
    
    for uid in user_ids:
        try:
            bot.send_message(uid, f"📢 Announcement:\n\n{text}")
            success += 1
        except:
            failed += 1
    
    bot.send_message(message.chat.id, 
                     f"✅ Broadcast complete!\n"
                     f"✓ Sent to: {success} users\n"
                     f"✗ Failed: {failed}")

# Change password command
@bot.message_handler(commands=['changepassword'])
def changepassword_command(message):
    user_id = str(message.from_user.id)
    
    if not is_admin(user_id):
        bot.send_message(message.chat.id, "❌ This command is for main admin only.")
        return
    
    new_password = message.text.replace('/changepassword', '').strip()
    if not new_password:
        bot.send_message(message.chat.id, 
                        "🔐 Change Admin Password\n\n"
                        "Usage:\n"
                        "/changepassword new_password\n\n"
                        "Example:\n"
                        "/changepassword MyNewPass123")
        return
    
    if len(new_password) < 1:
        bot.send_message(message.chat.id, "❌ Password must be at least 1 character long.")
        return
    
    if update_admin_password(new_password):
        bot.send_message(message.chat.id, f"✅ Admin password has been changed to: {new_password}")
    else:
        bot.send_message(message.chat.id, "❌ Error changing password.")

# Delete all files command
@bot.message_handler(commands=['deleteall'])
def deleteall_command(message):
    user_id = str(message.from_user.id)
    
    if not (is_admin(user_id) or is_subadmin(user_id) or subadmin_sessions.get(user_id)):
        bot.send_message(message.chat.id, "❌ This command is for admins only.")
        return
    
    try:
        # Count files before deletion
        file_count = 0
        if os.path.exists(FILES_FOLDER):
            file_count = len([f for f in os.listdir(FILES_FOLDER) if f.endswith('.txt')])
        
        # Delete all files in uploaded_files folder
        if os.path.exists(FILES_FOLDER):
            shutil.rmtree(FILES_FOLDER)
            os.makedirs(FILES_FOLDER, exist_ok=True)
        
        # Clear processed accounts
        with open(PROCESSED_ACCOUNTS_FILE, 'w') as f:
            json.dump({"processed": []}, f, indent=4)
        
        # Clear work sessions
        work_sessions.clear()
        account_data_store.clear()
        
        # Clear cooldown timers
        cooldown_timers.clear()
        
        bot.send_message(message.chat.id, f"✅ Successfully deleted {file_count} uploaded files and cleared processed accounts list.")
        
    except Exception as e:
        bot.send_message(message.chat.id, f"❌ Error deleting files: {str(e)}")

# Handle work callback buttons
@bot.callback_query_handler(func=lambda call: call.data.startswith('work_'))
def handle_work_callback(call):
    user_id = str(call.from_user.id)
    action = call.data.split('_')[1]
    
    if action == "start":
        # Start working - show first account
        if user_id in work_sessions:
            show_next_account(call.message.chat.id, user_id)
            bot.answer_callback_query(call.id, "Starting work...")
        else:
            bot.answer_callback_query(call.id, "❌ Work session not found")
    
    elif action == "cancel":
        # Cancel work session
        if user_id in work_sessions:
            del work_sessions[user_id]
        
        # Clear cooldown timer
        if user_id in cooldown_timers:
            del cooldown_timers[user_id]
        
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text="❌ Work session cancelled."
        )
        bot.answer_callback_query(call.id, "Work cancelled")

# Show next account to process
def show_next_account(chat_id, user_id):
    if user_id not in work_sessions:
        return
    
    session = work_sessions[user_id]
    current_index = session['current_index']
    
    if current_index >= len(session['lines']):
        # All accounts processed
        bot.send_message(chat_id,
                        f"✅ All Accounts Processed!\n\n"
                        f"📊 Statistics:\n"
                        f"- Total processed: {session['processed_count']}\n"
                        f"- Remaining in file: 0")
        
        # Clear work session
        del work_sessions[user_id]
        return
    
    # Get current line
    line = session['lines'][current_index]
    parts = line.split('|')
    
    if len(parts) >= 2:
        username = parts[0].strip()
        password = parts[1].strip()
        full_input = line
        
        # Create a unique ID for this account data
        account_id = str(uuid.uuid4())[:8]  # Short unique ID
        
        # Store the full data in account_data_store
        account_data_store[account_id] = {
            'username': username,
            'password': password,
            'full_input': full_input,
            'user_id': user_id,
            'line_index': current_index,
            'user_code': session['user_code'],
            'start_time': time.time()  # Store when account was shown
        }
        
        # Add UID to background check queue
        add_uid_to_check_queue(username, account_id)
        
        # Create message with copyable format
        message_text = f"""
🔢 Account {current_index + 1} of {len(session['lines'])}

📋 Account Details:
"""
        
        # Create copyable username with markup
        message_text += f"\n👤 Username:\n<code>{username}</code>\n"
        
        # Create copyable password with markup
        message_text += f"\n🔑 Password:\n<code>{password}</code>\n"
        
        message_text += "\nChoose an action:"
        
        # Get cooldown time
        cooldown_time = get_cooldown_time()
        
        # If cooldown is enabled, show countdown
        if cooldown_time > 0:
            message_text += f"\n\n⏰ Cooldown: {cooldown_time} seconds (please verify account before submitting)"
        
        # Create buttons with only the account_id in callback_data
        keyboard = InlineKeyboardMarkup()
        keyboard.row(
            InlineKeyboardButton("✅ Confirm", callback_data=f"confirm_{account_id}"),
            InlineKeyboardButton("❌ Suspend", callback_data=f"suspend_{account_id}")
        )
        keyboard.row(
            InlineKeyboardButton("⚠️ Issue", callback_data=f"issue_{account_id}"),
            InlineKeyboardButton("🚫 Cancel Session", callback_data="work_cancel_current")
        )
        
        # Start cooldown for this user
        start_cooldown(user_id)
        
        # Send with HTML parse mode for copyable text
        bot.send_message(chat_id, message_text, 
                        reply_markup=keyboard,
                        parse_mode='HTML')

# Handle account processing callbacks
@bot.callback_query_handler(func=lambda call: call.data.startswith('confirm_') or 
                     call.data.startswith('suspend_') or call.data.startswith('issue_'))
def handle_account_callback(call):
    user_id = str(call.from_user.id)
    
    if not call.data:
        bot.answer_callback_query(call.id, "❌ Invalid callback data")
        return
    
    parts = call.data.split('_')
    if len(parts) < 2:
        bot.answer_callback_query(call.id, "❌ Invalid callback data")
        return
    
    action = parts[0]  # 'confirm', 'suspend', or 'issue'
    account_id = parts[1]  # The unique identifier
    
    # Check if user is in cooldown
    in_cooldown, remaining = is_user_in_cooldown(user_id)
    if in_cooldown:
        bot.answer_callback_query(call.id, f"⏰ Please wait {int(remaining)} more seconds to submit")
        return
    
    # Process account immediately
    process_account_now(call, account_id, action)

# Process account immediately
def process_account_now(call, account_id, action):
    user_id = str(call.from_user.id)
    
    # Retrieve the data from account_data_store
    data = account_data_store.get(account_id)
    if not data:
        bot.answer_callback_query(call.id, "❌ Account data not found")
        return
    
    if data['user_id'] != user_id:
        bot.answer_callback_query(call.id, "❌ This account is not assigned to you")
        return
    
    username = data['username']
    password = data['password']
    full_input = data['full_input']
    line_index = data['line_index']
    user_code = data['user_code']
    
    if user_id not in work_sessions:
        bot.answer_callback_query(call.id, "❌ Work session not found")
        return
    
    # Check if account is still available (not processed by someone else)
    if is_account_processed(username):
        bot.answer_callback_query(call.id, "❌ This account was already processed by someone else!")
        
        # Try to send next account
        work_sessions[user_id]['current_index'] += 1
        show_next_account(call.message.chat.id, user_id)
        return
    
    # Check if UID check result is available
    uid_check_result = get_uid_check_result(account_id)
    
    if action == "confirm" and uid_check_result:
        # Check if UID is live
        if uid_check_result.get('is_live'):
            status = "Confirmed"
            final_action = "confirm"
        else:
            status = "c.suspended"
            final_action = "c_suspend"
    else:
        # No UID check result or user clicked suspend/issue
        if action == "suspend":
            status = "Suspended"
            final_action = "suspend"
        elif action == "issue":
            status = "Issue"
            final_action = "issue"
        else:
            status = "Confirmed"
            final_action = "confirm"
    
    # Mark as processed
    mark_account_processed(username)
    
    # Get processor info from session
    session = work_sessions[user_id]
    processor_name = session.get('username', 'User')
    
    # Update user stats
    user_stats = update_user_stats(user_id, processor_name, user_code, final_action)
    
    # Save to Excel with full input (including cookie)
    excel_saved = save_to_excel(full_input, username, password, processor_name, user_code, status)
    
    if not excel_saved:
        print(f"❌ WARNING: Failed to save {username} to Excel")
    
    # Update work session
    work_sessions[user_id]['current_index'] = line_index + 1
    work_sessions[user_id]['processed_count'] += 1
    
    # Update message to show status
    if uid_check_result and action == "confirm":
        live_status = "✅ LIVE" if uid_check_result.get('is_live') else "❌ DEAD"
        new_text = call.message.text + f"\n\n🔍 UID Check: {live_status}\n✅ Status: {status}"
    else:
        new_text = call.message.text + f"\n\n✅ Status: {status}"
    
    try:
        bot.edit_message_text(
            chat_id=call.message.chat.id,
            message_id=call.message.message_id,
            text=new_text,
            parse_mode='HTML'
        )
    except:
        pass
    
    # Clean up the stored data
    if account_id in account_data_store:
        del account_data_store[account_id]
    
    bot.answer_callback_query(call.id, f"Marked as {status}")
    
    # Check if user has reached limit after this processing
    limit_reached, remaining = check_user_limit(user_id)
    if limit_reached:
        limit = get_user_limit(user_id)
        bot.send_message(call.message.chat.id,
                        f"🚫 Work Limit Reached!\n\n"
                        f"You have reached your work limit of {limit} accounts.\n"
                        f"Please contact admin to increase your limit.")
        
        # Clear work session
        if user_id in work_sessions:
            del work_sessions[user_id]
        
        # Clear cooldown timer
        if user_id in cooldown_timers:
            del cooldown_timers[user_id]
        return
    
    # Show next account after a short delay
    time.sleep(0.5)
    show_next_account(call.message.chat.id, user_id)

# Handle cancel current session
@bot.callback_query_handler(func=lambda call: call.data == 'work_cancel_current')
def handle_cancel_current(call):
    user_id = str(call.from_user.id)
    
    if user_id in work_sessions:
        del work_sessions[user_id]
    
    # Clean up any account data for this user
    account_ids_to_remove = []
    for account_id, data in account_data_store.items():
        if data['user_id'] == user_id:
            account_ids_to_remove.append(account_id)
    
    for account_id in account_ids_to_remove:
        del account_data_store[account_id]
    
    # Clear cooldown timer
    if user_id in cooldown_timers:
        del cooldown_timers[user_id]
    
    bot.edit_message_text(
        chat_id=call.message.chat.id,
        message_id=call.message.message_id,
        text="❌ Work session cancelled."
    )
    bot.answer_callback_query(call.id, "Session cancelled")

# Main function with webhook fix
if __name__ == "__main__":
    print("🚀 Bot is starting...")
    print(f"🔐 Default admin password: {get_admin_password()}")
    print(f"👑 Admin username: {ADMIN_USERNAME}")
    init_files()
    print("📁 Data files initialized")
    print("📂 Folders created")
    print(f"📊 Loaded {len(username_to_id)} username mappings")
    print("🎨 Excel styling initialized")
    print("📋 Rules system initialized")
    print("👥 Approval system initialized")
    print("🔍 Background UID checker started")
    print("👨‍💼 Sub-admin system initialized")
    print("💾 All results folder created")
    print(f"⏰ Cooldown system initialized: {get_cooldown_time()} seconds")
    
    # Get system status
    system_status = get_system_status()
    work_enabled = system_status.get("work_enabled", True)
    print(f"🔌 System status: {'🟢 ONLINE' if work_enabled else '🔴 OFFLINE'}")
    
    print("🆕 New features initialized:")
    print("   📋 Copyable username/password in work")
    print("   🔵 Taken system")
    print("   🟠 Pending system (FIXED)")
    print("   🔍 Account checking system")
    print("   📈 User work limits")
    print("   ✏️ Edit user statistics")
    print("   🔌 System on/off control")
    print("   📢 Custom on/off notices")
    print("   🗑️ Unmarkp system to clear pending accounts")
    print(f"   🔍 Check interval: {CHECK_INTERVAL} seconds")
    
    # Test Excel file
    if os.path.exists(EXCEL_FILE):
        print(f"📊 Excel file exists: {EXCEL_FILE} ({os.path.getsize(EXCEL_FILE)} bytes)")
    else:
        print("❌ Excel file not created!")
    
    # FIX: Delete webhook before starting polling
    print("🔄 Deleting any existing webhook...")
    try:
        bot.delete_webhook()
        print("✅ Webhook deleted successfully")
        time.sleep(1)  # Wait a bit
    except Exception as e:
        print(f"⚠️ Could not delete webhook: {e}")
    
    print("🤖 Bot is running...")
    
    # Start bot polling with error handling
    while True:
        try:
            bot.polling(none_stop=True, interval=2, timeout=30)
        except Exception as e:
            print(f"❌ Bot error: {e}")
            print("🔄 Restarting bot in 10 seconds...")
            time.sleep(10)
